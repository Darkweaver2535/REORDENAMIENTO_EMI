#!/usr/bin/env python
"""
preview_excel_labs.py — Análisis pre-importación de archivos Excel de laboratorios
===================================================================================
Parser v3 con carry-forward y mapeo robusto de columnas.

Estructura confirmada por inspección directa:
  CBBA   → cabecera fila #6, col: 2=lab, 3=tipo, 4=sec, 5=sup, 6=ubi, 7..=asigs
  TRÓPICO → cabecera fila #5, col: 1=lab, 2=tipo, 3=sec, 4=sup, 5=ubi, 6..=asigs

Lógica carry-forward:
  - lab_actual  : se actualiza con el último NOMBRE DEL LABORATORIO no vacío
  - tipo_actual : se actualiza con el último tipo válido visto (para filas sin tipo)

Clasificación de filas:
  SUBESPACIO  → tiene sec (nombre_sec) → se registra con padre=lab_actual
  RAIZ        → tiene lab pero no sec → actualiza carry, no genera hijo
  ASIGNATURA  → solo columnas ≥ col_asig_start con datos útiles → descartada
  DESCRIPTIVA → fila de normas, sub-cabecera PEA/INVESTIGACIÓN → descartada
  AMBIGUA     → tiene sec pero sin padre o problema de tipo

No modifica la base de datos.
"""

import os
import sys
import argparse
from decimal import Decimal, InvalidOperation
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ARCHIVOS = {
    "COCHABAMBA (U.A. CBBA)": os.path.join(
        BASE_DIR,
        "COCHABAMBA-SISTEMA DE GESTIÓN DE LABORATORIOS 2026 U.A. CBBA.xlsx",
    ),
    "TRÓPICO": os.path.join(
        BASE_DIR,
        "TROPICO-SISTEMA DE GESTIÓN DE LABORATORIOS.xlsx",
    ),
}

# Subtipo canonical map
TIPO_CANON = {
    "SALA":        "SALA",
    "ÁREA":        "AREA",
    "AREA":        "AREA",
    "SECCIÓN":     "SECCION",
    "SECCION":     "SECCION",
    "LABORATORIO": "LABORATORIO",
}

# Palabras que identifican sub-cabeceras a descartar
SUBCAB_KEYWORDS = {"PEA", "INVESTIGACIÓN", "INVESTIGACION", "VENTA DE SERVICIOS"}


def norm(v) -> str:
    """Normaliza una celda a string limpio, sin saltos, sin nbsp."""
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").replace("\n", " ").strip()


def norm_upper(v) -> str:
    return norm(v).upper()


def canon_tipo(raw: str) -> str | None:
    """Retorna subtipo canónico o None si no reconocido."""
    u = norm_upper(raw).rstrip()
    # Buscar coincidencia exacta primero
    if u in TIPO_CANON:
        return TIPO_CANON[u]
    # Búsqueda por prefijo (ej: "SALA " → "SALA")
    for k, v in TIPO_CANON.items():
        if u.startswith(k):
            return v
    return None


def detectar_schema(ws):
    """
    Detecta la fila de cabecera y los índices de columna.
    Retorna (cab_fila_num, col_nom, col_tipo, col_sec, col_sup, col_ubi, col_asig_start)
    """
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), start=1):
        for j, cell in enumerate(row):
            if cell and "NOMBRE DEL LABORATORIO" in norm_upper(str(cell)):
                # Encontrada la cabecera — mapear todas las columnas de esta fila
                col_nom = col_tipo = col_sec = col_sup = col_ubi = None
                col_asig_start = 999

                for k, c in enumerate(row):
                    nu = norm_upper(str(c)) if c else ""
                    if not nu:
                        continue
                    if "NOMBRE DEL LABORATORIO" in nu and col_nom is None:
                        col_nom = k
                    elif ("SALA" in nu or "AREA" in nu or "ÁREA" in nu or "SECCIÓN" in nu) and "SELECCIONE" in nu and col_tipo is None:
                        col_tipo = k
                    elif "NOMBRE DE LA" in nu and col_sec is None:
                        col_sec = k
                    elif "SUPERFICIE" in nu and col_sup is None:
                        col_sup = k
                    elif ("UBICACIÓN" in nu or "UBICACION" in nu) and col_ubi is None:
                        col_ubi = k
                    elif any(kw in nu for kw in ("ASIGNATURA", "SEMESTRE", "CARRERA", "ACTIVIDAD", "NORMA")):
                        col_asig_start = min(col_asig_start, k)

                if col_asig_start == 999:
                    col_asig_start = (col_ubi or col_sup or col_sec or 5) + 1

                return i, col_nom, col_tipo, col_sec, col_sup, col_ubi, col_asig_start

    return None, None, None, None, None, None, None


def get_cell(row, col):
    if col is None or col >= len(row):
        return ""
    return norm(row[col])


def analizar_archivo(nombre, ruta, solo_errores=False):
    W = 72
    SEP = "═" * W
    sep = "─" * W

    print(f"\n{SEP}")
    print(f"  📂  {nombre}")
    print(f"  {os.path.basename(ruta)}")
    print(SEP)

    if not os.path.exists(ruta):
        print("  ❌ Archivo no encontrado.")
        return

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return

    ws = wb.active
    print(f"  Hoja: {ws.title!r}  |  Dims: {ws.dimensions}")

    cab, c_nom, c_tipo, c_sec, c_sup, c_ubi, c_asig = detectar_schema(ws)
    if cab is None:
        print("  ❌ No se encontró la cabecera.")
        return

    print(f"  Cabecera fila #{cab}")
    print(f"  Columnas → nom={c_nom} | tipo={c_tipo} | sec={c_sec} | sup={c_sup} | ubi={c_ubi} | asigs_desde={c_asig}")

    # ── Procesamiento fila a fila ────────────────────────────────────────────
    generales    = {}     # {lab: {sec: (tipo, sup_decimal, ubi)}}
    tipo_cntr    = defaultdict(int)
    ambiguas     = []     # (fila, motivo, detalle)
    descartadas  = []     # (fila, motivo, detalle)

    lab_actual  = None
    tipo_actual = None   # carry del último tipo válido

    all_rows = list(ws.iter_rows(min_row=cab + 1, values_only=True))

    for idx, row in enumerate(all_rows):
        fila_num = cab + 1 + idx

        nom  = get_cell(row, c_nom)
        tipo = get_cell(row, c_tipo)
        sec  = get_cell(row, c_sec)
        sup  = get_cell(row, c_sup)
        ubi  = get_cell(row, c_ubi)

        nom_u  = nom.upper()
        tipo_u = tipo.upper()
        sec_u  = sec.upper()

        # Celdas con contenido real (ignorando nbsp sueltos)
        non_empty = [j for j, c in enumerate(row)
                     if c is not None and norm(str(c))]

        # ── Fila totalmente vacía
        if not non_empty:
            continue

        # ── Celdas útiles (en la zona de lab/tipo/sec/sup/ubi)
        useful = [j for j in non_empty if j <= (c_asig - 1 if c_asig else 99)]

        # ── Sub-cabecera de actividades (PEA / INVESTIGACIÓN …)
        first_vals = [norm_upper(str(row[j])) for j in non_empty[:3]]
        if any(any(kw in fv for kw in SUBCAB_KEYWORDS) for fv in first_vals):
            descartadas.append((fila_num, "sub-cabecera actividades", str(first_vals[:2])[:50]))
            continue

        # ── Fila solo-asignatura: todos los datos están ≥ col_asig_start
        if not useful:
            descartadas.append((fila_num, "solo columnas de asignatura/norma", f"cols={non_empty[:5]}"))
            continue

        # ── Actualizar carry de lab si viene en esta fila
        if nom_u and nom_u not in {"N°", "NOMBRE DEL LABORATORIO", ""}:
            lab_actual = nom.strip()

        # ── Canonizar tipo de la celda
        tipo_canon = canon_tipo(tipo_u) if tipo_u else None
        if tipo_canon:
            tipo_actual = tipo_canon  # carry

        # ── Fila sin nombre de subespacio → es raíz/continuación
        if not sec.strip():
            if nom.strip():
                descartadas.append((fila_num, "fila-raíz: tiene lab pero sin sec",
                                    f"lab={nom.strip()!r:.45}"))
            else:
                descartadas.append((fila_num, "fila sin sec ni lab (sólo tipo?)",
                                    f"tipo={tipo!r} cols_utiles={useful[:4]}"))
            continue

        # ── Tiene nombre de subespacio → es un hijo
        # Necesitamos padre
        if not lab_actual:
            ambiguas.append((fila_num, "hijo sin padre carry",
                             f"sec={sec.strip()!r:.40} tipo={tipo!r}"))
            continue

        # Tipo: celda ó carry ó fallback LABORATORIO (con nota ambigua)
        tipo_final = tipo_canon or tipo_actual
        if not tipo_final:
            ambiguas.append((fila_num, "tipo no detectado (carry vacío → LABORATORIO)",
                             f"lab={lab_actual!r:.35} sec={sec.strip()!r:.30}"))
            tipo_final = "LABORATORIO"

        # Superficie
        sup_val = None
        if sup:
            try:
                sup_val = Decimal(sup.replace(",", ".").replace(" ", ""))
            except InvalidOperation:
                ambiguas.append((fila_num, f"SUPERFICIE no numérica: {sup!r}",
                                 f"sec={sec.strip()!r:.40}"))

        # ── Registrar (de-dup: mismo padre+mismo hijo = mismo subespacio)
        if lab_actual not in generales:
            generales[lab_actual] = {}

        sec_key = sec.strip()
        if sec_key not in generales[lab_actual]:
            generales[lab_actual][sec_key] = (tipo_final, sup_val, ubi)
            tipo_cntr[tipo_final] += 1

    # ── Estadísticas ─────────────────────────────────────────────────────────
    total_g   = len(generales)
    total_h   = sum(len(v) for v in generales.values())
    con_h     = sum(1 for v in generales.values() if v)
    sin_h     = total_g - con_h

    print(f"\n  {sep}")
    print(f"  📊  RESUMEN")
    print(f"  {sep}")
    print(f"  Nodos raíz  (GENERAL)     : {total_g}")
    print(f"    ├ Con subespacios        : {con_h}")
    print(f"    └ Solo raíz (sin hijos)  : {sin_h}")
    print(f"  Subespacios detectados    : {total_h}")
    print(f"  Relaciones raíz→hijo      : {total_h}")
    print(f"  Filas descartadas         : {len(descartadas)}")
    print(f"  Filas ambiguas            : {len(ambiguas)}")
    print()
    print(f"  Distribución por subtipo:")
    for t in ["SALA", "AREA", "SECCION", "LABORATORIO"]:
        n = tipo_cntr.get(t, 0)
        print(f"    {t:<12} {n:>3}  {'█'*n}")

    # ── Raíces y subespacios ─────────────────────────────────────────────────
    print(f"\n  {sep}")
    print(f"  🌳  RAÍCES Y SUBESPACIOS DETECTADOS")
    print(f"  {sep}")
    for lab, hijos in sorted(generales.items()):
        print(f"\n  ▸ {lab}  [{len(hijos)} subespacio(s)]")
        for hijo, (tipo, sup_d, ubi) in sorted(hijos.items()):
            sup_s = f"{sup_d}m²" if sup_d else "—"
            ubi_s = (ubi[:48] + "…") if ubi and len(ubi) > 48 else (ubi or "—")
            print(f"      [{tipo:<11}] {hijo:<38} {sup_s:<12} {ubi_s}")

    # ── Ambiguas y descartadas (primeras 10) ────────────────────────────────
    print(f"\n  {sep}")
    print(f"  ⚠  FILAS AMBIGUAS O DESCARTADAS — primeras 10")
    print(f"  {sep}")
    todas = [(f, m, d, "AMBIGUA ") for f, m, d in ambiguas] + \
            [(f, m, d, "DESCART.") for f, m, d in descartadas]
    todas.sort(key=lambda x: x[0])
    for f, m, d, cat in todas[:10]:
        print(f"  F{f:>3} [{cat}] {m:<47}  {d[:32]}")
    if len(todas) > 10:
        print(f"  … y {len(todas)-10} filas más omitidas en este listado")
    print()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--solo-errores", action="store_true")
    args = parser.parse_args()

    W = 72
    print("\n" + "═"*W)
    print("  SGL-EMI — Preview Importación Excel · Parser v3 (carry-forward)")
    print("  No modifica la base de datos.")
    print("═"*W)

    for nombre, ruta in ARCHIVOS.items():
        analizar_archivo(nombre, ruta, solo_errores=args.solo_errores)

    print("═"*W)
    print("  Para importar en real → añadir soporte a import_labs.py")
    print("  o renombrar columnas del Excel a: NOMBRE_LAB_GENERAL,")
    print("  TIPO_SECCION, NOMBRE_SECCION, SUPERFICIE, UBICACION")
    print("═"*W + "\n")


if __name__ == "__main__":
    main()
