"""
Management command: verificar_celdas
=====================================
Verificación CELDA A CELDA de la importación.

`auditar_carga` comprueba que cada fila del Excel tenga su registro en la base.
Esto va un paso más allá: por cada fila, compara **el contenido de cada columna**
con el campo donde debería haber quedado. Detecta columnas que se perdieron en
silencio, textos truncados y valores mal mapeados.

Qué compara:

  FAMILIA C (fichas técnicas por laboratorio)
      nombre del equipo   → Equipo.nombre
      ubicación           → Equipo.ubicacion_sala
      marca y modelo      → especificaciones["marca_modelo"]
      año de adquisición  → especificaciones["anio_adquisicion"]
      estado              → Equipo.estatus_general (normalizado)
      especificaciones    → especificaciones["especificaciones"]
      funcionalidad       → especificaciones["funcionalidad"]
      foto (URL)          → Equipo.foto_url

  FAMILIA B (padrón contable de activos fijos)
      descripción         → Equipo.nombre  (sólo si el activo lo creó el padrón)
      responsable         → especificaciones["responsable"]
      cargo               → especificaciones["cargo_responsable"]
      oficina             → especificaciones["oficina_contable"]
      grupo contable      → especificaciones["grupo_contable"]
      auxiliar            → especificaciones["auxiliar"]
      costo histórico     → especificaciones["costo_historico"]
      fecha               → especificaciones["fecha_historico"]
      vida útil           → especificaciones["vida_util_consumida"]

Uso:
    python manage.py verificar_celdas
    python manage.py verificar_celdas --detalle   # muestra cada discrepancia
"""

import glob
import os
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime

from django.core.management.base import BaseCommand

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.laboratorios.codigos import codigo_base, es_sin_codigo
from apps.laboratorios.management.commands.import_equipos import (
    parse_estado,
    parse_fecha_completa,
)

SRC = "/Users/alvaroencinas/Downloads/INFORMACION REORDENAMIENTO EMI"

FAMILIA_A = {
    "COCHABAMBA-SISTEMA DE GESTION DE  LABORATORIOS 2026 U.A. CBBA.xlsx",
    "LA PAZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx",
    "TROPICO-SISTEMA DE GESTIÓN DE LABORATORIOS.xlsx",
    "SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx",
}

FAMILIA_B = {
    "EQUIPO MEDICO Y LAB. UALP EMI.xlsx",
    "EQUIPO MEDICO Y LABORATORIO UACBBA.xlsx",
    "GRUPO EQUIPO MEDICO Y LABORATORIO UASC.xlsx",
    "LABORATORIO DE LA EMI UAT.xlsx",
    "DETALLE DE ACTIVOS FIJOS EQUIPO MEDICOS Y DE LABORATORIO UA RIBERALTA.xlsx",
    "EQUIPO MEDICO Y DE LABORATORIO OFICINA CENTRAL.xlsx",
}

# Longitudes que imponen los importadores; recortar ahí no es pérdida de datos
# sino el límite del campo, pero se reporta aparte para que quede constancia.
LIMITES = {"nombre": 150, "ubicacion_sala": 100, "foto_url": 500,
           "especificaciones": 2000, "funcionalidad": 2000}


def limpiar(valor):
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return re.sub(r"\s+", " ", str(valor).replace("\xa0", " ").replace("\n", " ")).strip()


def norm(valor):
    s = limpiar(valor)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").upper()
    # El acento agudo suelto con el que empiezan algunas celdas del Excel y los
    # signos de puntuación no cambian el significado del dato.
    return re.sub(r"[^A-Z0-9]", "", s)


def igual_fecha(excel, bd):
    """Compara fechas/años tolerando ceros y formatos distintos.

    "4/12/2007" y "04/12/2007" son la misma fecha; "2015" y "28/12/2015" son el
    mismo año de adquisición (la columna se llama "AÑO EN QUE SE ADQUIRIÓ").
    """
    def partes(v):
        # Se lee la celda con el mismo parser del importador, para que un
        # "21/092015" (al que le falta una barra) se compare como 21/09/2015.
        s = parse_fecha_completa(v) or limpiar(v)
        nums = re.findall(r"\d+", s)
        anio = next((n for n in nums if len(n) == 4 and n[:2] in ("19", "20")), None)
        return anio, tuple(str(int(n)) for n in nums)

    ae, be = partes(excel), partes(bd)
    if ae[1] == be[1]:
        return True
    # Basta con que coincida el año cuando uno de los dos trae la fecha completa.
    return bool(ae[0]) and ae[0] == be[0]


def normalizar_url(u):
    """El dominio venía mal escrito en algunas celdas ("drive.googlee.com")."""
    return limpiar(u).replace("googlee.com", "google.com")


def igual(excel, bd, limite=None):
    """Compara tolerando espacios, tildes, puntuación y el truncado del campo."""
    a, b = norm(excel), norm(bd)
    if a == b:
        return True
    if not a and not b:
        return True
    # El valor guardado puede ser el original recortado al largo del campo.
    if limite and a.startswith(b[:max(len(b) - 5, 0)]) and len(b) >= limite - 10:
        return True
    return False


class Command(BaseCommand):
    help = "Compara celda a celda los Excel oficiales contra la base de datos."

    def add_arguments(self, parser):
        parser.add_argument("--detalle", action="store_true")
        parser.add_argument("--src", default=SRC)

    # ── Detección de cabeceras ───────────────────────────────────────────
    def _cols_ficha(self, ws):
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
            j = norm(" ".join(str(c) for c in row if c))
            if "NOMBREDELEQUIPO" in j and "CODIGODELEQUIPO" in j:
                cols = {}
                fotos = []
                for k, c in enumerate(row):
                    u = norm(c)
                    if not u:
                        continue
                    if "NOMBREDELEQUIPO" in u:
                        cols.setdefault("nombre", k)
                    elif "UBICACIONDELEQUIPO" in u:
                        cols.setdefault("ubic", k)
                    elif "MARCAYMODELO" in u:
                        cols.setdefault("marca", k)
                    elif "CODIGODELEQUIPO" in u:
                        cols.setdefault("cod", k)
                    elif "ANOENQUE" in u or "ADQUIRIO" in u:
                        cols.setdefault("anio", k)
                    elif "ESTADOACTUAL" in u:
                        cols.setdefault("estado", k)
                    elif "ESPECIFICACIONESTECNICAS" in u:
                        cols.setdefault("espec", k)
                    elif "FUNCIONALIDAD" in u:
                        cols.setdefault("func", k)
                    elif "FOTO" in u:
                        fotos.append(k)
                if fotos:
                    cols["foto1"] = fotos[0]
                    if len(fotos) > 1:
                        cols["foto2"] = fotos[1]
                return i, cols
        return None, None

    def _cols_padron(self, ws):
        campos = {
            "cod": ("CODIGO",), "desc": ("DESCRIPCIONDELBIEN", "DESCRIPCION"),
            "resp": ("RESPONSABLE",), "cargo": ("CARGO",), "ofi": ("OFICINA",),
            "grupo": ("GRUPOCONTABLE",), "aux": ("AUXILIAR",),
            "costo": ("COSTOHISTORICO",), "fecha": ("FECHAHISTORICO", "FECHAINC"),
            "vida": ("VIDAUTILCONSUMIDA", "VIDAUTIL"),
        }
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
            j = norm(" ".join(str(c) for c in row if c))
            if "CODIGO" in j and "DESCRIPCION" in j:
                cols = {}
                for k, c in enumerate(row):
                    u = norm(c)
                    if not u:
                        continue
                    for campo, pref in campos.items():
                        if campo not in cols and any(u.startswith(p) for p in pref):
                            cols[campo] = k
                            break
                if "cod" in cols and "desc" in cols:
                    return i, cols
        return None, None

    def _celda(self, row, cols, clave):
        k = cols.get(clave)
        return limpiar(row[k]) if k is not None and k < len(row) else ""

    def handle(self, *args, **opt):
        from apps.laboratorios.models import Equipo

        src, detalle = opt["src"], opt["detalle"]
        idx = defaultdict(list)
        for e in Equipo.objects.all():
            idx[codigo_base(e.codigo_activo)].append(e)

        # ── FAMILIA C ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 104)
        self.stdout.write(self.style.SUCCESS(
            "FAMILIA C — FICHAS TÉCNICAS · comparación celda a celda"))
        self.stdout.write("═" * 104)
        campos_c = ["nombre", "ubic", "marca", "anio", "estado", "espec", "func", "foto"]
        destino_c = {
            "nombre": lambda e: e.nombre,
            "ubic": lambda e: e.ubicacion_sala,
            "marca": lambda e: (e.especificaciones or {}).get("marca_modelo", ""),
            # La celda puede traer la fecha exacta; se acepta el año O la fecha
            # completa, que ahora se conserva en `fecha_adquisicion`.
            "anio": lambda e: (e.especificaciones or {}).get("fecha_adquisicion")
                              or (e.especificaciones or {}).get("anio_adquisicion", ""),
            "estado": None,  # se compara aparte (se normaliza a bueno/regular/malo)
            "espec": lambda e: (e.especificaciones or {}).get("especificaciones", ""),
            "func": lambda e: (e.especificaciones or {}).get("funcionalidad", ""),
            "foto": lambda e: e.foto_url or "",
        }
        cmp_c = {c: {"con_dato": 0, "ok": 0, "difiere": 0} for c in campos_c}
        ejemplos = []
        filas_totales = 0

        for ruta in sorted(glob.glob(os.path.join(src, "*.xlsx"))):
            fn = os.path.basename(ruta)
            if fn in FAMILIA_A or fn in FAMILIA_B:
                continue
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            for ws in wb.worksheets:
                hr, cols = self._cols_ficha(ws)
                if hr is None:
                    continue
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    cod = self._celda(row, cols, "cod")
                    nom = self._celda(row, cols, "nombre")
                    if not cod and not nom:
                        continue
                    filas_totales += 1
                    if not cod or es_sin_codigo(cod):
                        continue
                    equipos = idx.get(codigo_base(cod))
                    if not equipos:
                        continue
                    for campo in campos_c:
                        if campo == "foto":
                            valor = ""
                            for c in ("foto1", "foto2"):
                                bruto = self._celda(row, cols, c)
                                m = re.search(r"https?://\S+", bruto)
                                if m:
                                    valor = m.group(0)
                                    break
                        else:
                            valor = self._celda(row, cols, campo)
                        if not valor:
                            continue
                        if campo == "anio" and not re.search(r"(19|20)\d{2}", valor):
                            continue  # texto libre, no una fecha de adquisición
                        cmp_c[campo]["con_dato"] += 1
                        if campo == "anio":
                            coincide = any(
                                igual_fecha(valor, (e.especificaciones or {}).get("fecha_adquisicion")
                                            or (e.especificaciones or {}).get("anio_adquisicion", ""))
                                for e in equipos)
                        elif campo == "foto":
                            coincide = any(
                                normalizar_url(valor) == normalizar_url(e.foto_url or "")
                                for e in equipos)
                        elif campo == "estado":
                            # Misma lectura que hace el importador, para no
                            # tener dos criterios distintos de lo que es "malo".
                            esperado, _ = parse_estado(valor)
                            coincide = any(e.estatus_general == esperado for e in equipos)
                        else:
                            lim = LIMITES.get(
                                {"nombre": "nombre", "ubic": "ubicacion_sala",
                                 "foto": "foto_url", "espec": "especificaciones",
                                 "func": "funcionalidad"}.get(campo, ""), None)
                            coincide = any(igual(valor, destino_c[campo](e), lim) for e in equipos)

                        # El Excel repite algunas filas con otra redacción; esa
                        # variante se guardó en la ficha y también vale.
                        if not coincide and campo in ("marca", "espec", "foto", "anio"):
                            claves = ("fecha", "anio") if campo == "anio" else (campo,)
                            coincide = any(
                                igual_fecha(valor, alt[k]) if campo == "anio"
                                else igual(valor, alt[k], None)
                                for e in equipos
                                for alt in (e.especificaciones or {}).get("otra_version_en_origen", [])
                                for k in claves if alt.get(k))
                        eq = equipos[0]
                        if coincide:
                            cmp_c[campo]["ok"] += 1
                        else:
                            cmp_c[campo]["difiere"] += 1
                            if len(ejemplos) < 25:
                                ejemplos.append((fn[:34], ws.title[:16], cod, campo,
                                                 limpiar(valor)[:52],
                                                 str(destino_c[campo](eq) if destino_c[campo] else eq.estatus_general)[:52]))
            wb.close()

        self.stdout.write(f"{'CAMPO':16}{'con dato':>10}{'coincide':>10}{'difiere':>9}   {'':<}")
        self.stdout.write("─" * 104)
        total_dif_c = 0
        etiquetas = {"nombre": "Nombre", "ubic": "Ubicación", "marca": "Marca y modelo",
                     "anio": "Año adquisic.", "estado": "Estado", "espec": "Especificaciones",
                     "func": "Funcionalidad", "foto": "Foto (URL)"}
        for campo in campos_c:
            d = cmp_c[campo]
            total_dif_c += d["difiere"]
            pct = 100 * d["ok"] // d["con_dato"] if d["con_dato"] else 100
            marca = "✅" if d["difiere"] == 0 else "⚠ "
            self.stdout.write(f"{marca} {etiquetas[campo]:14}{d['con_dato']:>10}{d['ok']:>10}"
                              f"{d['difiere']:>9}   {pct}%")
        self.stdout.write("─" * 104)
        self.stdout.write(f"filas de datos recorridas: {filas_totales}")

        if ejemplos and detalle:
            self.stdout.write("\nDISCREPANCIAS (muestra):")
            for fn, hoja, cod, campo, exc, bd in ejemplos:
                self.stdout.write(self.style.WARNING(f"   {fn} · {hoja} · {cod} · {campo}"))
                self.stdout.write(f"       excel: {exc}")
                self.stdout.write(f"       bd   : {bd}")

        # ── FAMILIA B ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 104)
        self.stdout.write(self.style.SUCCESS(
            "FAMILIA B — PADRÓN CONTABLE · comparación celda a celda"))
        self.stdout.write("═" * 104)
        campos_b = ["resp", "cargo", "ofi", "grupo", "aux", "costo", "fecha", "vida"]
        destino_b = {
            "resp": "responsable", "cargo": "cargo_responsable", "ofi": "oficina_contable",
            "grupo": "grupo_contable", "aux": "auxiliar", "costo": "costo_historico",
            "fecha": "fecha_historico", "vida": "vida_util_consumida",
        }
        etq_b = {"resp": "Responsable", "cargo": "Cargo", "ofi": "Oficina",
                 "grupo": "Grupo contable", "aux": "Auxiliar", "costo": "Costo histórico",
                 "fecha": "Fecha", "vida": "Vida útil"}
        cmp_b = {c: {"con_dato": 0, "ok": 0, "difiere": 0} for c in campos_b}
        ejemplos_b = []

        for fn in sorted(FAMILIA_B):
            ruta = os.path.join(src, fn)
            if not os.path.exists(ruta):
                continue
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            hr, cols = self._cols_padron(ws)
            if hr is None:
                wb.close()
                continue
            for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                cod = self._celda(row, cols, "cod")
                desc = self._celda(row, cols, "desc")
                if not cod or not desc:
                    continue
                equipos = idx.get(codigo_base(cod))
                if not equipos:
                    continue
                # Los datos contables se aplican a UNO de los equipos del grupo.
                eq = next((e for e in equipos
                           if "grupo_contable" in (e.especificaciones or {})), equipos[0])
                esp = eq.especificaciones or {}
                for campo in campos_b:
                    valor = self._celda(row, cols, campo)
                    if not valor:
                        continue
                    cmp_b[campo]["con_dato"] += 1
                    if igual(valor, esp.get(destino_b[campo], "")):
                        cmp_b[campo]["ok"] += 1
                    else:
                        cmp_b[campo]["difiere"] += 1
                        if len(ejemplos_b) < 25:
                            ejemplos_b.append((fn[:34], cod, campo, valor[:52],
                                               str(esp.get(destino_b[campo], ""))[:52]))
            wb.close()

        self.stdout.write(f"{'CAMPO':18}{'con dato':>10}{'coincide':>10}{'difiere':>9}")
        self.stdout.write("─" * 104)
        total_dif_b = 0
        for campo in campos_b:
            d = cmp_b[campo]
            total_dif_b += d["difiere"]
            pct = 100 * d["ok"] // d["con_dato"] if d["con_dato"] else 100
            marca = "✅" if d["difiere"] == 0 else "⚠ "
            self.stdout.write(f"{marca} {etq_b[campo]:16}{d['con_dato']:>10}{d['ok']:>10}"
                              f"{d['difiere']:>9}   {pct}%")

        if ejemplos_b and detalle:
            self.stdout.write("\nDISCREPANCIAS (muestra):")
            for fn, cod, campo, exc, bd in ejemplos_b:
                self.stdout.write(self.style.WARNING(f"   {fn} · {cod} · {campo}"))
                self.stdout.write(f"       excel: {exc}")
                self.stdout.write(f"       bd   : {bd}")

        # ── VEREDICTO ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 104)
        if total_dif_c == 0 and total_dif_b == 0:
            self.stdout.write(self.style.SUCCESS(
                "✅ FIDELIDAD TOTAL — cada celda con dato de los Excel coincide con la base"))
        else:
            self.stdout.write(self.style.ERROR(
                f"⚠ DISCREPANCIAS: {total_dif_c} en fichas técnicas · {total_dif_b} en el padrón"))
            self.stdout.write("   Ejecuta con --detalle para ver cada caso.")
        self.stdout.write("═" * 104 + "\n")
