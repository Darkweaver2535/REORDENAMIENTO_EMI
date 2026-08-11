"""
Management command: import_labs
================================
Importa laboratorios jerárquicos desde un Excel (.xlsx) de la EMI.

Lógica de parser (v3, carry-forward, metadatos académicos):
  - Auto-detecta la fila de cabecera buscando "NOMBRE DEL LABORATORIO"
  - Arrastra el último lab general no vacío como padre (lab_actual)
  - Arrastra el último tipo válido (tipo_actual)
  - Arrastra el último nodo afectado (ultimo_nodo_afectado) para vincular usos académicos
  - CRITERIO ESTRUCTURAL LITERAL: Los usos académicos (Asignaturas) se asignan
    estrictamente al subespacio definido en la misma fila del Excel, independientemente
    de si el nombre de la asignatura sugiere semánticamente otra dependencia.
  - Extrae y acumula banderas de actividades (PEA, Investigación, Servicios) y Normativa.
  - Crea registros en UsoAcademico para las asignaturas asociadas.

Clasifica cada fila:
  CREATE_ROOT      → lab nuevo sin hijo
  CREATE_CHILD     → subespacio nuevo
  UPDATE_CHILD     → subespacio actualizado
  EXIST_ROOT       → lab general existía
  EXIST_CHILD      → subespacio existía
  ADD_USO_ACAD     → fila sin estructura pero con asignatura/actividad
  SKIP_NORM_ROW    → fila de sub-cabecera
  WARN_MISMATCH    → asig no coincide semánticamente con su nodo padre
  AMBIGUOUS        → sec presente pero sin padre

Reimportar es seguro: los nodos y los usos que ya existen se reconocen y no se
duplican. La única excepción son los tres ambientes de electrónica que la
planilla de La Paz lista dos veces, bajo dos raíces distintas; el importador es
literal y los vuelve a crear tal como están escritos. Por eso el orden de la
carga es:

    import_labs  →  fusionar_duplicados

`fusionar_duplicados` los une de nuevo y es idempotente, así que puede correrse
siempre después de cualquier reimportación.
"""

import re
import unicodedata
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.laboratorios.management.commands.seed_estructura import clave_uso_academico

try:
    import openpyxl
except ImportError:
    openpyxl = None


A_CREATE_ROOT = "CREATE_ROOT"
A_CREATE_CHILD = "CREATE_CHILD"
A_UPDATE_CHILD = "UPDATE_CHILD"
A_EXIST_ROOT = "EXIST_ROOT"
A_EXIST_CHILD = "EXIST_CHILD"
A_ADD_USO_ACAD = "ADD_USO_ACAD"
A_SKIP_NORM_ROW = "SKIP_NORM_ROW"
A_WARN_MISMATCH = "WARN_MISMATCH"
A_AMBIGUOUS = "AMBIGUOUS"
A_FIX_NOMBRE_GEN = "FIX_NOMBRE_GEN"

# Nombres de subespacio sin significado propio (la celda repite el tipo de
# espacio); se componen con el nombre del laboratorio padre al importar.
NOMBRES_GENERICOS = {"LABORATORIO", "LABORATORIOS", "SALA", "AREA", "SECCION", "LAB"}

TIPO_CANON = {
    "SALA": "SALA",
    "AREA": "AREA",
    "SECCION": "SECCION",
    "LABORATORIO": "LABORATORIO",
    # Variantes fuera de catálogo vistas en los Excel de sedes
    "HUERTA": "AREA",
}

SUBCAB_KEYWORDS = {"PEA", "INVESTIGACION", "VENTA DE SERVICIOS"}


def normalizar_texto(valor: str) -> str:
    if not valor:
        return ""
    s = str(valor).replace("\xa0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper()


def limpiar_celda(valor) -> str:
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor).replace("\xa0", " ").replace("\n", " ")).strip()


def canon_tipo(raw: str) -> str | None:
    n = normalizar_texto(raw)
    if n in TIPO_CANON:
        return TIPO_CANON[n]
    for k, v in TIPO_CANON.items():
        if n.startswith(k):
            return v
    return None


def buscar_nodo(Laboratorio, nombre: str, unidad_academica, parent=None):
    """Busca un nodo existente ignorando tildes, mayúsculas y espacios dobles.

    `nombre__iexact` de Django SÍ distingue tildes, así que el mismo laboratorio
    escrito "ELECTRONICA" en el Excel de la sede y "ELECTRÓNICA" en el
    consolidado se duplicaba. Comparamos sobre el nombre normalizado entre los
    hermanos candidatos (mismo padre y misma unidad).
    """
    objetivo = normalizar_texto(nombre)
    if not objetivo:
        return None
    qs = Laboratorio.objects.filter(unidad_academica=unidad_academica)
    qs = qs.filter(parent__isnull=True) if parent is None else qs.filter(parent=parent)
    for cand in qs:
        if normalizar_texto(cand.nombre) == objetivo:
            return cand
    return None


def detectar_schema(ws):
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), start=1):
        for cell in row:
            if cell and "NOMBRE DEL LABORATORIO" in normalizar_texto(str(cell)):
                col_nom = col_tipo = col_sec = col_sup = col_ubi = None
                col_asig = col_sem = col_car = col_pea = col_inv = col_ven = col_nor = None
                col_asig_start = 999

                for k, c in enumerate(row):
                    nu = normalizar_texto(str(c)) if c else ""
                    if not nu:
                        continue
                    if "NOMBRE DEL LABORATORIO" in nu and col_nom is None:
                        col_nom = k
                    elif (
                        ("SALA" in nu or "AREA" in nu or "SECCION" in nu)
                        and "SELECCIONE" in nu
                        and col_tipo is None
                    ):
                        col_tipo = k
                    elif "NOMBRE DE LA" in nu and col_sec is None:
                        col_sec = k
                    elif "SUPERFICIE" in nu and col_sup is None:
                        col_sup = k
                    elif ("UBICACION" in nu or "UBICACIÓN" in nu) and col_ubi is None:
                        col_ubi = k
                    elif "ASIGNATURA" in nu and col_asig is None:
                        col_asig = k
                        col_asig_start = min(col_asig_start, k)
                    elif "SEMESTRE" in nu and col_sem is None:
                        col_sem = k
                    elif "CARRERA" in nu and col_car is None:
                        col_car = k
                    elif "ACTIVIDAD" in nu and col_pea is None:
                        col_pea = k
                        col_inv = k + 1
                        col_ven = k + 2
                    elif "NORMA" in nu and col_nor is None:
                        col_nor = k

                if col_asig_start == 999:
                    col_asig_start = (col_ubi or col_sup or col_sec or 5) + 1

                return (
                    i,
                    col_nom,
                    col_tipo,
                    col_sec,
                    col_sup,
                    col_ubi,
                    col_asig_start,
                    col_asig,
                    col_sem,
                    col_car,
                    col_pea,
                    col_inv,
                    col_ven,
                    col_nor,
                )
    return (None,) * 14


def get_cell(row, col) -> str:
    if col is None or col >= len(row):
        return ""
    return limpiar_celda(row[col])


_COLORS = {
    A_CREATE_ROOT: "\033[92m",
    A_CREATE_CHILD: "\033[32m",
    A_UPDATE_CHILD: "\033[33m",
    A_EXIST_ROOT: "\033[90m",
    A_EXIST_CHILD: "\033[90m",
    A_ADD_USO_ACAD: "\033[36m",
    A_SKIP_NORM_ROW: "\033[90m",
    A_WARN_MISMATCH: "\033[35m",
    A_AMBIGUOUS: "\033[91m",
    A_FIX_NOMBRE_GEN: "\033[93m",
}
_RESET = "\033[0m"


def color_action(action: str) -> str:
    return f"{_COLORS.get(action, '')}{action:<17}{_RESET}"


class Command(BaseCommand):
    help = "Importa laboratorios y usos académicos desde un Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)
        parser.add_argument("--unidad-academica", required=True, metavar="UA")
        parser.add_argument("--campus", default="", metavar="CAMPUS")
        parser.add_argument(
            "--hoja",
            default=None,
            metavar="HOJA",
            help="Nombre de la hoja a importar (por defecto, la hoja activa del workbook).",
        )
        parser.add_argument(
            "--alias",
            default=None,
            metavar="JSON",
            help=(
                "JSON {nombre_en_el_excel: nombre_canónico}. Sirve para que un "
                "consolidado que llama distinto al mismo espacio no cree un nodo "
                "duplicado (ej. 'PLANTA DE AGUAS' → 'PLANTA DE AGUAS Y ALIMENTOS')."
            ),
        )
        parser.add_argument(
            "--solo-nuevos",
            action="store_true",
            help=(
                "Ignora las filas cuyo laboratorio raíz ya existe en la BD. "
                "Sirve para complementar con un consolidado sin pisar el detalle "
                "que ya cargó el archivo de la sede."
            ),
        )
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--verbose", action="store_true")

    def _resolver_unidad(self, UnidadAcademica, ref: str):
        if ref.strip().isdigit():
            return UnidadAcademica.objects.get(pk=int(ref))
        ref_n = normalizar_texto(ref)
        for ua in UnidadAcademica.objects.all():
            if normalizar_texto(ua.nombre) == ref_n:
                return ua
        candidatos = [
            ua for ua in UnidadAcademica.objects.all() if ref_n in normalizar_texto(ua.nombre)
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        raise CommandError(f"Referencia '{ref}' ambigua/no encontrada.")

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl no instalado")

        from apps.estructura_academica.models import UnidadAcademica
        from apps.laboratorios.models import Laboratorio, UsoAcademico

        archivo, ua_ref, campus, dry_run, verbose = (
            options["archivo"],
            options["unidad_academica"],
            options["campus"],
            options["dry_run"],
            options["verbose"],
        )
        solo_nuevos = options["solo_nuevos"]
        # Alias: nombre tal como viene en el Excel → nombre canónico ya en la BD.
        #
        # La clave admite un ámbito, porque `normalizar_texto` ignora tildes y un
        # alias pensado para un subespacio puede colisionar con el nombre de su
        # propio padre (el consolidado tiene la raíz "EDAFOLOGIA" con el hijo
        # "EDAFOLOGÍA": ambos normalizan igual):
        #     "hijo:X"  → sólo subespacios
        #     "raiz:X"  → sólo laboratorios raíz
        #     "X"       → ambos
        alias_raiz, alias_hijo = {}, {}
        if options["alias"]:
            import json as _json
            with open(options["alias"], encoding="utf-8") as fh:
                crudo = _json.load(fh)
            for clave, valor in crudo.items():
                if clave.startswith("_"):
                    continue
                ambito, _, nombre = clave.partition(":")
                ambito = ambito.lower()
                if not nombre:
                    ambito, nombre = "", clave
                n = normalizar_texto(nombre)
                if ambito in ("", "raiz"):
                    alias_raiz[n] = valor
                if ambito in ("", "hijo"):
                    alias_hijo[n] = valor

        def canonizar_raiz(nombre):
            return alias_raiz.get(normalizar_texto(nombre), nombre)

        def canonizar_hijo(nombre):
            return alias_hijo.get(normalizar_texto(nombre), nombre)

        if dry_run:
            self.stdout.write(self.style.WARNING("\n⚠ DRY-RUN activado\n"))
        unidad_academica = self._resolver_unidad(UnidadAcademica, ua_ref)
        self.stdout.write(f"📚 UA: {unidad_academica}\n")

        wb = openpyxl.load_workbook(archivo, data_only=True)
        hoja = options["hoja"]
        if hoja:
            if hoja not in wb.sheetnames:
                raise CommandError(
                    f"Hoja '{hoja}' no existe en el archivo. Disponibles: {wb.sheetnames}"
                )
            ws = wb[hoja]
        else:
            ws = wb.active
        self.stdout.write(f"📄 Hoja: {ws.title}\n")

        schema = detectar_schema(ws)
        if schema[0] is None:
            raise CommandError("No se encontró cabecera")
        (
            cab,
            c_nom,
            c_tipo,
            c_sec,
            c_sup,
            c_ubi,
            c_asig_start,
            c_asig,
            c_sem,
            c_car,
            c_pea,
            c_inv,
            c_ven,
            c_nor,
        ) = schema

        cnts = {k: 0 for k in _COLORS.keys()}

        lab_actual = None
        lab_nom_disp = None
        tipo_actual = None
        ultimo_nodo_afectado = None
        # Se recalcula en cada cambio de laboratorio raíz; con --solo-nuevos marca
        # que las filas siguientes pertenecen a un lab ya cargado y hay que saltarlas.
        saltar_raiz = False

        def log(fila_num, action, msg=""):
            if verbose:
                self.stdout.write(f"  F{fila_num:>3} {color_action(action)} {msg}")
            cnts[action] += 1

        with transaction.atomic():
            for idx, row in enumerate(ws.iter_rows(min_row=cab + 1, values_only=True)):
                fila_num = cab + 1 + idx

                nom = get_cell(row, c_nom)
                tipo = get_cell(row, c_tipo)
                sec = get_cell(row, c_sec)
                sup = get_cell(row, c_sup)
                ubi = get_cell(row, c_ubi)

                asig = get_cell(row, c_asig)
                sem = get_cell(row, c_sem)
                car = get_cell(row, c_car)
                pea = get_cell(row, c_pea)
                inv = get_cell(row, c_inv)
                ven = get_cell(row, c_ven)
                nor = get_cell(row, c_nor)

                nom_n = normalizar_texto(nom)

                non_empty = [
                    j for j, c in enumerate(row) if c is not None and limpiar_celda(str(c))
                ]
                if not non_empty:
                    continue

                # Sub-cabeceras
                # Igualdad exacta: con substring, asignaturas como "METODOLOGÍA DE LA
                # INVESTIGACIÓN" se descartaban como sub-cabecera.
                first_vals = [normalizar_texto(str(row[j])) for j in non_empty[:3]]
                if any(fv in SUBCAB_KEYWORDS for fv in first_vals):
                    log(fila_num, A_SKIP_NORM_ROW, "sub-cabecera")
                    continue

                action_main = None

                if nom_n and nom_n not in {"N", "NOMBRE DEL LABORATORIO"}:
                    nuevo_lab_nom = canonizar_raiz(nom.strip())
                    if nuevo_lab_nom != lab_nom_disp:
                        lab_nom_disp = nuevo_lab_nom
                        if not dry_run:
                            lab_obj = buscar_nodo(
                                Laboratorio, lab_nom_disp, unidad_academica, parent=None
                            )
                            created = lab_obj is None
                            if created:
                                lab_obj = Laboratorio.objects.create(
                                    nombre=lab_nom_disp,
                                    clase_nodo=Laboratorio.ClaseNodo.GENERAL,
                                    unidad_academica=unidad_academica,
                                    campus=campus,
                                )
                            lab_actual = lab_obj
                            ultimo_nodo_afectado = lab_obj
                            action_main = A_CREATE_ROOT if created else A_EXIST_ROOT
                            # --solo-nuevos: el laboratorio ya lo cargó una fuente de
                            # mayor prioridad; no tocamos su detalle.
                            saltar_raiz = solo_nuevos and not created
                        else:
                            action_main = A_CREATE_ROOT
                            ultimo_nodo_afectado = "MOCK_ROOT"
                            saltar_raiz = False

                        if action_main:
                            log(fila_num, action_main, f"lab={lab_nom_disp!r}")

                if saltar_raiz:
                    continue

                tipo_canon = canon_tipo(tipo) if tipo else None
                if tipo_canon:
                    tipo_actual = tipo_canon

                # Proceso de subespacio
                if sec.strip():
                    if not lab_nom_disp:
                        log(fila_num, A_AMBIGUOUS, f"sec={sec.strip()!r} sin padre")
                        continue

                    # Nombre genérico (la celda repite el tipo, ej. "LABORATORIO"):
                    # se compone con el padre para que el nodo tenga significado.
                    if normalizar_texto(sec) in NOMBRES_GENERICOS:
                        sec_original = sec.strip()
                        sec = f"{sec_original} DE {lab_nom_disp}"
                        log(fila_num, A_FIX_NOMBRE_GEN, f"{sec_original!r} → {sec!r}")

                    sec = canonizar_hijo(sec.strip())
                    tipo_final = tipo_canon or tipo_actual or "LABORATORIO"
                    sup_val = None
                    if sup:
                        # Tolera unidades y formatos sucios: "57.20 m²", "59,61 m2", "78, 10 m3"
                        m = re.search(r"\d+(?:\s*[.,]\s*\d+)?", sup)
                        if m:
                            try:
                                sup_val = Decimal(m.group().replace(",", ".").replace(" ", ""))
                            except InvalidOperation:
                                pass

                    if not dry_run:
                        # La identidad del nodo es (nombre, padre, unidad) — igual
                        # que la restricción `uq_lab_nombre_parent_ua`. El subtipo
                        # NO entra en la búsqueda: el mismo subespacio aparece como
                        # ÁREA en el Excel de la sede y como LABORATORIO en el
                        # consolidado, y buscarlo por subtipo lo duplicaría.
                        hijo = buscar_nodo(
                            Laboratorio, sec.strip(), unidad_academica, parent=lab_actual
                        )
                        created = hijo is None
                        if created:
                            hijo = Laboratorio.objects.create(
                                nombre=sec.strip(),
                                parent=lab_actual,
                                unidad_academica=unidad_academica,
                                campus=campus,
                                clase_nodo=Laboratorio.ClaseNodo.SUBESPACIO,
                                subtipo_espacio=tipo_final,
                                superficie_m2=sup_val,
                                ubicacion=ubi or "",
                            )
                        ultimo_nodo_afectado = hijo

                        # Al complementar con un consolidado sólo se RELLENAN los
                        # huecos: no se pisa la superficie ni la ubicación que ya
                        # trajo el Excel de la sede, que es la fuente autoritativa.
                        changed = False
                        if sup_val is not None and hijo.superficie_m2 is None:
                            hijo.superficie_m2 = sup_val
                            changed = True
                        if ubi and not hijo.ubicacion:
                            hijo.ubicacion = ubi
                            changed = True
                        if changed:
                            hijo.save(update_fields=["superficie_m2", "ubicacion"])
                            action_main = A_UPDATE_CHILD
                        else:
                            action_main = A_CREATE_CHILD if created else A_EXIST_CHILD
                    else:
                        ultimo_nodo_afectado = "MOCK_CHILD"
                        action_main = A_CREATE_CHILD

                    log(fila_num, action_main, f"[{tipo_final}] {sec.strip()!r}")

                # Proceso de Usos Académicos y Banderas
                has_acad = asig or pea or inv or ven or nor
                if has_acad and ultimo_nodo_afectado:
                    if not action_main:
                        log(fila_num, A_ADD_USO_ACAD, f"Asignatura: {asig!r}")

                    if (
                        asig
                        and ultimo_nodo_afectado
                        and getattr(ultimo_nodo_afectado, "nombre", None)
                    ):
                        n_nodo = normalizar_texto(ultimo_nodo_afectado.nombre)
                        n_asig = normalizar_texto(asig)
                        # Comprobar heurísticamente si el nodo parece ajeno al nombre de la asignatura
                        if len(n_nodo) > 4 and n_nodo not in n_asig and "LABORATORIO" not in n_nodo:
                            # Puede existir desajuste semántico
                            log(
                                fila_num,
                                A_WARN_MISMATCH,
                                f"'{asig}' -> asignado al nodo literal '{ultimo_nodo_afectado.nombre}'",
                            )

                    if not dry_run:
                        changed = False
                        # boolean flags (any non-empty text usually means "X" or "SI")
                        if pea and not ultimo_nodo_afectado.usa_pea:
                            ultimo_nodo_afectado.usa_pea = True
                            changed = True
                        if inv and not ultimo_nodo_afectado.usa_investigacion:
                            ultimo_nodo_afectado.usa_investigacion = True
                            changed = True
                        if ven and not ultimo_nodo_afectado.usa_venta_servicios:
                            ultimo_nodo_afectado.usa_venta_servicios = True
                            changed = True
                        if nor:
                            actual_norma = ultimo_nodo_afectado.normativa_infraestructura or ""
                            if nor not in actual_norma:
                                sep = " | " if actual_norma else ""
                                ultimo_nodo_afectado.normativa_infraestructura = (
                                    actual_norma + sep + nor
                                )
                                changed = True
                        if changed:
                            ultimo_nodo_afectado.save(
                                update_fields=[
                                    "usa_pea",
                                    "usa_investigacion",
                                    "usa_venta_servicios",
                                    "normativa_infraestructura",
                                ]
                            )

                        if asig:
                            # El uso académico se identifica por el trío
                            # (asignatura, semestre, carrera), no sólo por la
                            # asignatura: la misma materia se dicta para varias
                            # carreras. Buscando sólo por el nombre se perdían
                            # esas filas y, al reimportar, `get_or_create`
                            # reventaba con MultipleObjectsReturned al
                            # encontrar dos usos con la misma asignatura.
                            clave = clave_uso_academico(asig, sem, car)
                            existe = any(
                                clave_uso_academico(u.asignatura, u.semestre, u.carrera) == clave
                                for u in UsoAcademico.objects.filter(
                                    laboratorio=ultimo_nodo_afectado)
                            )
                            if not existe:
                                UsoAcademico.objects.create(
                                    laboratorio=ultimo_nodo_afectado,
                                    asignatura=asig,
                                    semestre=sem,
                                    carrera=car,
                                )

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write("\n" + "─" * 62)
        self.stdout.write(self.style.SUCCESS(f"📊 RESUMEN{' (DRY-RUN)' if dry_run else ''}:"))
        for action, count in cnts.items():
            color = _COLORS.get(action, "")
            self.stdout.write(f"   {color}{action:<20}{_RESET}  {count}")
        self.stdout.write("─" * 62 + "\n")
