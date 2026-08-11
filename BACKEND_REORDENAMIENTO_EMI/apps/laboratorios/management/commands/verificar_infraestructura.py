"""
Management command: verificar_infraestructura
=============================================
Verificación celda a celda de la FAMILIA A (infraestructura de laboratorios).

`verificar_celdas` cubre las fichas de equipos y el padrón contable. Falta el
tercer documento: las planillas "INFRAESTRUCTURA DE LABORATORIOS" de cada sede,
que son las que describen el ambiente en sí. Este comando compara, por cada
laboratorio o subespacio de esas planillas:

    SUPERFICIE DEL AMBIENTE (m2)        → Laboratorio.superficie_m2
    UBICACIÓN                           → Laboratorio.ubicacion
    NORMA NACIONAL O INTERNACIONAL      → Laboratorio.normativa_infraestructura
    ACTIVIDAD · PEA                     → Laboratorio.usa_pea
    ACTIVIDAD · INVESTIGACIÓN           → Laboratorio.usa_investigacion
    ACTIVIDAD · VENTA DE SERVICIOS      → Laboratorio.usa_venta_servicios
    ASIGNATURA / SEMESTRE / CARRERA     → UsoAcademico (asignatura, semestre, carrera)

Las planillas usan celdas combinadas: el nombre del laboratorio aparece una sola
vez y las filas siguientes continúan el mismo ambiente. Se arrastra el contexto
igual que hace el importador, y por cada ambiente se toma el primer valor no
vacío de cada columna.

Uso:
    python manage.py verificar_infraestructura
    python manage.py verificar_infraestructura --detalle
"""

import os
import re
import unicodedata
from collections import defaultdict

from django.core.management.base import BaseCommand

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.laboratorios.management.commands.auditar_carga import (
    FAMILIA_A,
    NOMBRES_GENERICOS,
)
from apps.laboratorios.management.commands.auditar_carga import Command as Auditar
from apps.laboratorios.management.commands.auditar_carga import norm

SRC = "/Users/alvaroencinas/Downloads/INFORMACION REORDENAMIENTO EMI"

ACTIVIDADES = ("PEA", "INVESTIGACION", "VENTA DE SERVICIOS")

# Encabezados de las columnas de datos del ambiente.
CABECERAS = {
    "lab": "NOMBRE DEL LABORATORIO",
    "sub": "NOMBRE DE LA SALA",
    "sup": "SUPERFICIE DEL AMBIENTE",
    "ubic": "UBICACION",
    "asig": "QUE ASIGNATURAS UTILIZAN",
    "sem": "DE QUE SEMESTRE",
    "carr": "DE QUE CARRERA",
    "norma": "QUE NORMA NACIONAL",
}


def superficie(valor):
    """Devuelve los m² como número.

    Se lee con el mismo criterio que el importador: las celdas traen
    "57.20 m²", "360 m2", 89.84 y también "78, 10 m3", donde la coma es el
    separador decimal escrito con un espacio detrás (78,10 m²).
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    m = re.search(r"\d+(?:\s*[.,]\s*\d+)?", str(valor))
    if not m:
        return None
    return round(float(m.group(0).replace(",", ".").replace(" ", "")), 2)


def marcado(valor):
    """Las columnas de actividad se marcan con una X (o cualquier texto)."""
    return bool(norm(valor).strip())


def texto_igual(excel, bd):
    """Compara texto tolerando espacios, tildes, puntuación y viñetas."""
    def n(v):
        return re.sub(r"[^A-Z0-9]", "", norm(v))
    a, b = n(excel), n(bd)
    if a == b or (not a and not b):
        return True
    # El campo puede haberse guardado recortado a su longitud máxima.
    return bool(b) and len(b) >= 240 and a.startswith(b[:230])


class Command(BaseCommand):
    help = "Compara celda a celda las planillas de infraestructura contra la base."

    def add_arguments(self, parser):
        parser.add_argument("--detalle", action="store_true")
        parser.add_argument("--src", default=SRC)

    def _columnas(self, ws):
        """(fila_cabecera, {campo: col}, {actividad: col}) de la planilla."""
        for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True), 1):
            texto = norm(" ".join(str(c) for c in row if c))
            if CABECERAS["lab"] not in texto:
                continue
            cols = {}
            for k, c in enumerate(row):
                u = norm(c)
                if not u:
                    continue
                for campo, marca in CABECERAS.items():
                    if marca in u and campo not in cols:
                        cols[campo] = k
            # Las tres actividades viven en la fila siguiente (cabecera partida).
            act = {}
            siguiente = next(ws.iter_rows(min_row=i + 1, max_row=i + 1, values_only=True), ())
            for k, c in enumerate(siguiente):
                u = norm(c)
                for a in ACTIVIDADES:
                    if u == a and a not in act:
                        act[a] = k
            return i, cols, act
        return None, None, None

    def _leer(self, ruta, hoja, alias):
        """{nombre_normalizado: {campos…}} de un ambiente por fila del Excel."""
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb[hoja]
        hr, cols, act = self._columnas(ws)
        if hr is None:
            wb.close()
            return {}

        def celda(row, campo):
            k = cols.get(campo)
            return row[k] if k is not None and k < len(row) else None

        ambientes = {}
        raiz = nodo = ""
        for row in ws.iter_rows(min_row=hr + 2, values_only=True):
            t = norm(celda(row, "lab"))
            if t and t not in ACTIVIDADES:
                raiz = t
                nodo = norm(alias.get(("raiz", t), t))
            s = norm(celda(row, "sub"))
            if s and s not in ACTIVIDADES:
                if s in NOMBRES_GENERICOS and raiz:
                    s = f"{s} DE {raiz}"
                nodo = norm(alias.get(("hijo", s), s))
            if not nodo:
                continue

            amb = ambientes.setdefault((raiz, nodo), {
                "sup": None, "ubic": "", "norma": [], "usos": set(),
                "act": {a: False for a in ACTIVIDADES},
            })
            if amb["sup"] is None:
                amb["sup"] = superficie(celda(row, "sup"))
            if not amb["ubic"]:
                amb["ubic"] = re.sub(r"\s+", " ", str(celda(row, "ubic") or "")).strip()
            linea = re.sub(r"\s+", " ", str(celda(row, "norma") or "")).strip()
            if linea and linea not in amb["norma"]:
                amb["norma"].append(linea)
            for a, k in act.items():
                if k < len(row) and marcado(row[k]):
                    amb["act"][a] = True
            asig = norm(celda(row, "asig"))
            if asig and asig not in ACTIVIDADES:
                amb["usos"].add((asig, norm(celda(row, "sem")), norm(celda(row, "carr"))))
        wb.close()
        return ambientes

    def handle(self, *args, **opt):
        from apps.estructura_academica.models import UnidadAcademica
        from apps.laboratorios.models import Laboratorio, UsoAcademico

        detalle, src = opt["detalle"], opt["src"]
        auditor = Auditar()

        # Índice de laboratorios por (UA, nombre) y por (UA, padre, nombre).
        labs, por_padre = defaultdict(dict), defaultdict(dict)
        for lab in Laboratorio.objects.select_related("unidad_academica", "parent"):
            abrev = norm(getattr(lab.unidad_academica, "abreviacion", "") or "")
            clave = norm(lab.nombre)
            previo = labs[abrev].get(clave)
            # Varios laboratorios generales tienen un único subespacio con su
            # mismo nombre ("QUÍMICA" ⊃ "QUÍMICA"). El ambiente que describe la
            # planilla —superficie, ubicación, usos— es el subespacio.
            if previo is None or (previo.parent_id is None and lab.parent_id is not None):
                labs[abrev][clave] = lab
            padre = norm(lab.parent.nombre) if lab.parent else ""
            por_padre[abrev][(padre, clave)] = lab
        usos = defaultdict(set)
        for u in UsoAcademico.objects.all():
            usos[u.laboratorio_id].add((norm(u.asignatura), norm(u.semestre), norm(u.carrera)))

        campos = ["sup", "ubic", "norma", "pea", "inv", "venta", "usos"]
        cmp = {c: {"con_dato": 0, "ok": 0, "difiere": 0, "conflicto": 0} for c in campos}
        registros, fallos, conflictos, sin_nodo = [], [], [], []

        for fn, hoja, ua in FAMILIA_A:
            ruta = os.path.join(src, fn)
            if not os.path.exists(ruta):
                continue
            origen = "consolidado" if fn.startswith("SANTA CRUZ-DATOS") else "sede"
            ambientes = self._leer(ruta, hoja, auditor._alias(ua))
            indice = labs.get(norm(ua), {})
            bajo_padre = por_padre.get(norm(ua), {})
            for (raiz, nombre), amb in ambientes.items():
                # Primero por (padre, nombre); si la raíz del Excel no coincide
                # con el padre en la base (alias), se cae al nombre a secas.
                lab = bajo_padre.get((raiz, nombre)) or indice.get(nombre)
                if lab is None:
                    sin_nodo.append(f"{ua} · {nombre}")
                    continue

                def revisar(campo, valor_excel, coincide, mostrado_bd):
                    registros.append({
                        "lab": lab.pk, "campo": campo, "ok": coincide,
                        "fuente": f"{ua}/{origen}", "nombre": nombre,
                        "excel": str(valor_excel)[:58], "bd": str(mostrado_bd)[:58],
                    })

                if amb["sup"] is not None:
                    bd = float(lab.superficie_m2) if lab.superficie_m2 is not None else None
                    revisar("sup", amb["sup"], bd is not None and abs(bd - amb["sup"]) < 0.01, bd)
                if amb["ubic"]:
                    revisar("ubic", amb["ubic"],
                            texto_igual(amb["ubic"], lab.ubicacion), lab.ubicacion)
                if amb["norma"]:
                    # Cada línea del Excel debe estar dentro de lo guardado.
                    guardado = re.sub(r"[^A-Z0-9]", "", norm(lab.normativa_infraestructura))
                    revisar("norma", " | ".join(amb["norma"]),
                            all(re.sub(r"[^A-Z0-9]", "", norm(l)) in guardado
                                for l in amb["norma"]),
                            lab.normativa_infraestructura)
                for campo, clave, attr in (("pea", "PEA", "usa_pea"),
                                           ("inv", "INVESTIGACION", "usa_investigacion"),
                                           ("venta", "VENTA DE SERVICIOS", "usa_venta_servicios")):
                    if amb["act"][clave]:
                        revisar(campo, "X", getattr(lab, attr), getattr(lab, attr))
                for uso in amb["usos"]:
                    guardados = usos.get(lab.pk, set())
                    # El semestre y la carrera pueden venir en blanco en el Excel.
                    coincide = any(
                        g[0] == uso[0]
                        and (not uso[1] or g[1] == uso[1])
                        and (not uso[2] or g[2] == uso[2])
                        for g in guardados)
                    revisar("usos", " / ".join(x for x in uso if x),
                            coincide, f"{len(guardados)} usos registrados")

        # Campos que alguna fila logró cuadrar: lo demás son variantes de texto.
        cuadrados = {(r["lab"], r["campo"]) for r in registros if r["ok"]}
        for r in registros:
            d = cmp[r["campo"]]
            d["con_dato"] += 1
            if r["ok"]:
                d["ok"] += 1
            elif (r["lab"], r["campo"]) in cuadrados:
                d["conflicto"] += 1
                conflictos.append(r)
            else:
                d["difiere"] += 1
                fallos.append(r)

        etiquetas = {"sup": "Superficie (m²)", "ubic": "Ubicación",
                     "norma": "Normativa", "pea": "Actividad PEA",
                     "inv": "Actividad Investigación",
                     "venta": "Actividad Venta de servicios",
                     "usos": "Usos académicos"}
        ancho = 104
        self.stdout.write("\n" + "═" * ancho)
        self.stdout.write(self.style.SUCCESS(
            "FAMILIA A — INFRAESTRUCTURA DE LABORATORIOS · comparación celda a celda"))
        self.stdout.write("═" * ancho)
        self.stdout.write(f"{'CAMPO':32}{'con dato':>10}{'coincide':>10}"
                          f"{'otra redacción':>16}{'difiere':>9}")
        self.stdout.write("─" * ancho)
        total_dif = 0
        for c in campos:
            d = cmp[c]
            total_dif += d["difiere"]
            cubierto = d["ok"] + d["conflicto"]
            pct = f"{cubierto * 100 // d['con_dato']}%" if d["con_dato"] else "—"
            icono = "✅" if d["difiere"] == 0 else "⚠ "
            self.stdout.write(f"{icono} {etiquetas[c]:30}{d['con_dato']:>8}{d['ok']:>10}"
                              f"{d['conflicto']:>16}{d['difiere']:>9}   {pct}")
        self.stdout.write("─" * ancho)

        if sin_nodo:
            self.stdout.write(self.style.WARNING(
                f"\nAmbientes del Excel sin laboratorio en la base: {len(sin_nodo)}"))
            for s in sin_nodo[:15]:
                self.stdout.write(f"     {s}")

        def volcar(titulo, filas):
            self.stdout.write(f"\n{titulo}:")
            for r in filas[:60]:
                self.stdout.write(self.style.WARNING(
                    f"   {r['fuente']} · {r['nombre'][:34]} · {etiquetas[r['campo']]}"))
                self.stdout.write(f"       excel: {r['excel']}")
                self.stdout.write(f"       bd   : {r['bd']}")

        if conflictos and detalle:
            volcar("OTRA REDACCIÓN DEL MISMO DATO (la base guarda la de la sede)",
                   conflictos)
        if fallos and detalle:
            volcar("DISCREPANCIAS", fallos)

        self.stdout.write("\n" + "═" * ancho)
        if total_dif == 0 and not sin_nodo:
            self.stdout.write(self.style.SUCCESS(
                "✅ INFRAESTRUCTURA FIEL — cada celda con dato coincide con la base"))
        else:
            self.stdout.write(self.style.WARNING(
                f"⚠ DISCREPANCIAS: {total_dif} celdas · {len(sin_nodo)} ambientes sin nodo"))
            if not detalle:
                self.stdout.write("   Ejecuta con --detalle para ver cada caso.")
        self.stdout.write("═" * ancho + "\n")
