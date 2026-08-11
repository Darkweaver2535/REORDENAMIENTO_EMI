"""
Management command: import_equipos
==================================
Importa el inventario físico de equipos desde los Excel (.xlsx) de la EMI
(formato "DATOS DE EQUIPOS DE LABORATORIOS" — DNICYT / sedes).

Estructura esperada del Excel (una o varias hojas por archivo):
  - Fila(s) de título: "EQUIPOS DE LABORATORIO", "DATOS DE EQUIPOS ..."
  - Una celda "NOMBRE DEL LABORATORIO: <nombre>"  → identifica el lab destino
  - Fila de cabecera con columnas (posición variable, se auto-detecta):
        N° | NOMBRE DEL EQUIPO | UBICACIÓN DEL EQUIPO | MARCA Y MODELO |
        FOTO FRONTAL | FOTO | CÓDIGO DEL EQUIPO | AÑO | ESTADO ACTUAL |
        ESPECIFICACIONES TÉCNICAS | FUNCIONALIDAD
  - Filas de datos: cada fila = 1 unidad física (cantidad_total = 1).

Mapeo hoja → laboratorio:
  1. Mapeo explícito (--map archivo.json  ó  overrides internos).
  2. Coincidencia difusa (normalizada) contra los nodos HOJA de la UA.
  3. Si no hay coincidencia, se crea un nodo hoja nuevo (bajo el padre difuso
     si lo hay, o como nodo GENERAL con el nombre del laboratorio).

El estado ("BUENO" / "REGULAR" / "MALO" y variantes sucias como
"REGULAR, REQUIERE ...") se normaliza a la primera palabra clave y se reparte
en cantidad_buena / cantidad_regular / cantidad_mala (1 unidad).

Uso típico:
  # Un archivo (todas sus hojas):
  python manage.py import_equipos "ruta/archivo.xlsx" --unidad-academica UALP --dry-run
  # Un directorio completo:
  python manage.py import_equipos --dir "ruta/carpeta" --unidad-academica UACB --replace-ua --dry-run

Siempre revisa primero con --dry-run: imprime el plan hoja→lab y los conteos.
"""

import glob
import json
import os
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None


def normalizar(valor) -> str:
    if valor is None:
        return ""
    s = str(valor).replace("\xa0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper()


def limpiar(valor) -> str:
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor).replace("\xa0", " ").replace("\n", " ")).strip()


# Palabras del nombre de laboratorio que no aportan a la coincidencia difusa.
_STOP = {"DE", "DEL", "LA", "EL", "LOS", "LAS", "Y", "E", "LABORATORIO", "LAB",
         "AULA", "SALA", "AREA", "SECCION", "INGENIERIA", "UACB", "UALP", "UAT"}


def _tokens(nombre: str) -> set[str]:
    return {t for t in normalizar(nombre).split() if t and t not in _STOP}


def parse_estado(raw: str) -> tuple[str, str]:
    """Devuelve (estatus_general, texto_extra). estatus ∈ {bueno,regular,malo}."""
    n = normalizar(raw)
    if not n:
        return "regular", ""
    # Primera palabra clave (antes de coma/barra/dos puntos/punto)
    first = re.split(r"[,/:;.]", n, 1)[0].strip()
    extra = raw.strip()
    if first.startswith("BUEN") or first.startswith("NUEVO") or first.startswith("OPERA"):
        return "bueno", extra
    if first.startswith("REGULAR") or first.startswith("MEDIO"):
        return "regular", extra
    if (
        first.startswith("MAL")
        or first.startswith("NO ")
        or first == "NO"
        or first.startswith("CUMPLI")
        or first.startswith("PROVIENE")
        or first.startswith("DAN")
        or first.startswith("INSERV")
        or first.startswith("OBSOLET")
    ):
        return "malo", extra
    # Heurística de respaldo por contenido global
    if "BUEN" in n:
        return "bueno", extra
    if "MAL" in n or "INSERV" in n:
        return "malo", extra
    return "regular", extra


def parse_anio(raw: str) -> str:
    """Extrae un año de valores como '2016', '2016-12-28 00:00:00', datetime.

    Si la celda no contiene ningún año devuelve "": muchas fichas escriben ahí
    notas como "SIN REGISTRO" y guardarlas haría que la ficha del equipo
    mostrara «Año de adquisición: SIN REGISTRO». Esa nota se conserva aparte,
    en `especificaciones["nota_adquisicion"]`.
    """
    m = re.search(r"(19|20)\d{2}", limpiar(raw))
    return m.group(0) if m else ""


def parse_fecha_completa(raw) -> str:
    """Devuelve la fecha de adquisición completa cuando la celda la trae.

    La columna se llama "AÑO EN QUE SE ADQUIRIÓ", pero muchas fichas ponen la
    fecha exacta (14/08/2017). Guardar sólo el año descartaba el día y el mes.
    """
    from datetime import date, datetime as _dt

    if isinstance(raw, (_dt, date)):
        return raw.strftime("%d/%m/%Y")
    s = limpiar(raw)
    m = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-]((?:19|20)\d{2})\b", s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    m = re.search(r"\b((?:19|20)\d{2})[/-](\d{1,2})[/-](\d{1,2})\b", s)
    if m:
        return f"{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}"
    # A alguna celda le falta la barra del mes: "21/092015" es 21/09/2015.
    m = re.fullmatch(r"(\d{1,2})[/-](\d{2})((?:19|20)\d{2})", s)
    if m and 1 <= int(m.group(2)) <= 12:
        return f"{int(m.group(1)):02d}/{m.group(2)}/{m.group(3)}"
    return ""


def primer_url(*vals) -> str:
    for v in vals:
        s = limpiar(v)
        m = re.search(r"https?://\S+", s)
        if m:
            return m.group(0)[:500]
    return ""


class Command(BaseCommand):
    help = "Importa el inventario de equipos desde Excel(s) de laboratorio."

    def add_arguments(self, parser):
        parser.add_argument("archivo", nargs="?", default=None,
                            help="Ruta a un .xlsx (todas sus hojas).")
        parser.add_argument("--dir", default=None,
                            help="Directorio con múltiples .xlsx a importar.")
        parser.add_argument("--unidad-academica", required=True, metavar="UA")
        parser.add_argument("--campus", default="", metavar="CAMPUS")
        parser.add_argument("--map", default=None,
                            help="JSON {texto_hoja_o_lab: nombre_lab_destino}.")
        parser.add_argument("--map-filas", default=None,
                            help=(
                                "JSON {texto_columna_UBICACIÓN: nombre_lab_destino}. "
                                "Rutea fila por fila cuando una sola hoja mezcla varias "
                                "salas (ej. LABORATORIOS CIVIL trae hormigones, suelos, "
                                "asfaltos e hidráulica juntos). Si la fila no coincide, "
                                "cae al laboratorio resuelto para la hoja."
                            ))
        parser.add_argument("--replace-ua", action="store_true",
                            help="Elimina TODOS los equipos de la UA antes de importar.")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--verbose", action="store_true")

    # ── Resolución de UA ─────────────────────────────────────────────────────
    def _resolver_ua(self, UnidadAcademica, ref: str):
        if ref.strip().isdigit():
            return UnidadAcademica.objects.get(pk=int(ref))
        ref_n = normalizar(ref)
        for ua in UnidadAcademica.objects.all():
            if normalizar(ua.nombre) == ref_n:
                return ua
        cands = [ua for ua in UnidadAcademica.objects.all() if ref_n in normalizar(ua.nombre)]
        if len(cands) == 1:
            return cands[0]
        raise CommandError(f"UA '{ref}' ambigua/no encontrada.")

    # ── Resolución de laboratorio destino ────────────────────────────────────
    def _resolver_lab(self, Laboratorio, ua, lab_hint, sheet_name, fname, overrides, dry_run):
        """Devuelve (lab_obj_o_None, metodo, nombre_destino).

        Overrides (mayor a menor prioridad): 'archivo::hoja', 'archivo',
        'hoja', 'texto_del_hint'.
        """
        def _nfc(s):
            return unicodedata.normalize("NFC", s) if s else s

        ovr = (overrides.get(_nfc(f"{fname}::{sheet_name.strip()}"))
               or overrides.get(_nfc(fname))
               or overrides.get(_nfc(sheet_name.strip()))
               or overrides.get(_nfc(lab_hint.strip())))
        # Con override, ese texto es autoritativo (no se mezcla con hint/hoja).
        candidatos_texto = [ovr] if ovr else [t for t in (lab_hint, sheet_name) if t]
        candidatos_texto = candidatos_texto or [sheet_name]
        objetivo = candidatos_texto[0]

        labs = list(Laboratorio.objects.filter(unidad_academica=ua))
        hojas = [l for l in labs if l.es_hoja()]

        for texto in candidatos_texto:
            n = normalizar(texto)
            # 1) Igualdad exacta contra hoja
            for l in hojas:
                if normalizar(l.nombre) == n:
                    return l, "exacto", l.nombre

        # Con override explícito: no se fusiona por parecido. Si no hubo match
        # exacto, se crea un lab con ese nombre (intención autoritativa).
        # Sintaxis "PADRE>HIJO": crea HIJO como subespacio del lab PADRE existente.
        if ovr:
            padres = [l for l in labs if not l.es_hoja()]
            if ">" in objetivo:
                padre_txt, hijo_txt = (p.strip() for p in objetivo.split(">", 1))
                nombre_nuevo = hijo_txt
                padre_match = next(
                    (p for p in padres if normalizar(p.nombre) == normalizar(padre_txt)), None)
            else:
                nombre_nuevo = limpiar(objetivo)
                padre_match = next(
                    (p for p in padres if normalizar(p.nombre) == normalizar(nombre_nuevo)), None)
            if dry_run:
                m = f"CREAR hijo de «{padre_match.nombre}»" if padre_match else "CREAR (override)"
                return None, m, nombre_nuevo
            if padre_match:
                lab, _ = Laboratorio.objects.get_or_create(
                    nombre__iexact=nombre_nuevo, parent=padre_match,
                    defaults={"nombre": nombre_nuevo, "unidad_academica": ua,
                              "campus": padre_match.campus,
                              "clase_nodo": Laboratorio.ClaseNodo.SUBESPACIO,
                              "subtipo_espacio": Laboratorio.SubtipoEspacio.LABORATORIO})
                return lab, f"creado hijo de «{padre_match.nombre}»", lab.nombre
            lab, _ = Laboratorio.objects.get_or_create(
                nombre__iexact=nombre_nuevo, clase_nodo=Laboratorio.ClaseNodo.GENERAL,
                unidad_academica=ua, parent__isnull=True,
                defaults={"nombre": nombre_nuevo, "campus": "",
                          "clase_nodo": Laboratorio.ClaseNodo.GENERAL})
            return lab, "creado (override)", lab.nombre

        for texto in candidatos_texto:
            tks = _tokens(texto)
            if not tks:
                continue
            # 2) Subconjunto de tokens (evita QUIMICA⊂FISICOQUIMICA por substring)
            for l in hojas:
                lt = _tokens(l.nombre)
                if lt and (tks <= lt or lt <= tks):
                    return l, "tokens⊆", l.nombre
        # 3) Solape de tokens (mejor Jaccard sobre hojas)
        toks = _tokens(objetivo)
        mejor, mejor_score = None, 0.0
        if toks:
            for l in hojas:
                lt = _tokens(l.nombre)
                if not lt:
                    continue
                inter = len(toks & lt)
                if inter:
                    score = inter / len(toks | lt)
                    if score > mejor_score:
                        mejor, mejor_score = l, score
        if mejor and mejor_score >= 0.5:
            return mejor, f"tokens:{mejor_score:.2f}", mejor.nombre

        # 4) Sin coincidencia hoja → intentar padre para colgar un hijo nuevo
        padres = [l for l in labs if not l.es_hoja()]
        padre_match = None
        for texto in candidatos_texto:
            n = normalizar(texto)
            for p in padres:
                pn = normalizar(p.nombre)
                if pn == n or (len(n) > 3 and (n in pn or pn in n)):
                    padre_match = p
                    break
            if padre_match:
                break

        nombre_nuevo = limpiar(objetivo) or sheet_name.strip()
        if dry_run:
            metodo = f"CREAR hijo de «{padre_match.nombre}»" if padre_match else "CREAR raíz nueva"
            return None, metodo, nombre_nuevo

        if padre_match:
            lab, _ = Laboratorio.objects.get_or_create(
                nombre__iexact=nombre_nuevo,
                parent=padre_match,
                defaults={
                    "nombre": nombre_nuevo,
                    "unidad_academica": ua,
                    "campus": padre_match.campus,
                    "clase_nodo": Laboratorio.ClaseNodo.SUBESPACIO,
                    "subtipo_espacio": Laboratorio.SubtipoEspacio.LABORATORIO,
                },
            )
            return lab, f"creado hijo de «{padre_match.nombre}»", lab.nombre

        lab, _ = Laboratorio.objects.get_or_create(
            nombre__iexact=nombre_nuevo,
            clase_nodo=Laboratorio.ClaseNodo.GENERAL,
            unidad_academica=ua,
            parent__isnull=True,
            defaults={"nombre": nombre_nuevo, "campus": "",
                      "clase_nodo": Laboratorio.ClaseNodo.GENERAL},
        )
        return lab, "creado raíz", lab.nombre

    # ── Detección de cabecera / columnas ─────────────────────────────────────
    def _detectar(self, ws):
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
            joined = normalizar(" ".join(str(c) for c in row if c))
            if "NOMBRE DEL EQUIPO" in joined and "CODIGO DEL EQUIPO" in joined:
                cols = {}
                for k, c in enumerate(row):
                    u = normalizar(c)
                    if not u:
                        continue
                    if "NOMBRE DEL EQUIPO" in u and "nombre" not in cols:
                        cols["nombre"] = k
                    elif "UBICACION DEL EQUIPO" in u and "ubic" not in cols:
                        cols["ubic"] = k
                    elif "MARCA" in u and "marca" not in cols:
                        cols["marca"] = k
                    elif "FOTO" in u:
                        cols.setdefault("foto1", k)
                        if cols.get("foto1") != k:
                            cols.setdefault("foto2", k)
                        elif "foto2" not in cols and cols.get("foto1") == k:
                            pass
                    elif "CODIGO DEL EQUIPO" in u and "cod" not in cols:
                        cols["cod"] = k
                    elif ("AÑO" in str(c).upper() or "ADQUIRIO" in u or "ADQUIRIÓ" in str(c).upper()) and "anio" not in cols:
                        cols["anio"] = k
                    elif "ESTADO" in u and "estado" not in cols:
                        cols["estado"] = k
                    elif "ESPECIFICACION" in u and "espec" not in cols:
                        cols["espec"] = k
                    elif "FUNCIONALIDAD" in u and "func" not in cols:
                        cols["func"] = k
                # segunda columna FOTO
                foto_cols = [k for k, c in enumerate(row) if c and "FOTO" in normalizar(c)]
                if foto_cols:
                    cols["foto1"] = foto_cols[0]
                    if len(foto_cols) > 1:
                        cols["foto2"] = foto_cols[1]
                # nombre del laboratorio (rows arriba de la cabecera)
                lab_hint = ""
                for r2 in ws.iter_rows(min_row=1, max_row=i, values_only=True):
                    for c in r2:
                        if not c:
                            continue
                        m = re.search(r"NOMBRE DEL LABORATORIO\s*:?\s*(.+)", str(c), re.IGNORECASE)
                        if m and m.group(1).strip():
                            lab_hint = re.sub(r"\s+", " ", m.group(1)).strip()
                            break
                    if lab_hint:
                        break
                return i, cols, lab_hint
        return None, None, ""

    def _cell(self, row, cols, key):
        k = cols.get(key)
        if k is None or k >= len(row):
            return ""
        return limpiar(row[k])

    def handle(self, *args, **opt):
        if openpyxl is None:
            raise CommandError("openpyxl no instalado")

        from apps.estructura_academica.models import UnidadAcademica
        from apps.laboratorios.models import Equipo, Laboratorio

        ua = self._resolver_ua(UnidadAcademica, opt["unidad_academica"])
        dry = opt["dry_run"]
        verbose = opt["verbose"]
        overrides = {}
        if opt["map"]:
            with open(opt["map"], encoding="utf-8") as fh:
                raw_map = json.load(fh)
            overrides = {unicodedata.normalize("NFC", k): v
                         for k, v in raw_map.items() if not k.startswith("_")}

        # Ruteo por fila (columna UBICACIÓN → laboratorio), con las claves ya
        # normalizadas para que "ÁREA DE PESAJE" case con "AREA DE PESAJE".
        rutas_fila = {}
        if opt["map_filas"]:
            with open(opt["map_filas"], encoding="utf-8") as fh:
                rutas_fila = {normalizar(k): v
                              for k, v in json.load(fh).items() if not k.startswith("_")}
        cache_rutas = {}

        def resolver_por_ubicacion(ubic, lab_hoja):
            """Laboratorio destino de una fila según su columna UBICACIÓN."""
            destino = rutas_fila.get(normalizar(ubic))
            if not destino:
                return lab_hoja
            if destino in cache_rutas:
                return cache_rutas[destino]
            objetivo = normalizar(destino)
            lab = next(
                (l for l in Laboratorio.objects.filter(unidad_academica=ua)
                 if normalizar(l.nombre) == objetivo and l.es_hoja()),
                None,
            )
            if lab is None:
                self.stdout.write(self.style.WARNING(
                    f"   ⚠ --map-filas: no existe la hoja «{destino}» en {ua.nombre}; "
                    f"esas filas quedan en el lab de la hoja."))
                lab = lab_hoja
            cache_rutas[destino] = lab
            return lab

        # Lista de archivos
        archivos = []
        if opt["archivo"]:
            archivos.append(opt["archivo"])
        if opt["dir"]:
            archivos += sorted(glob.glob(os.path.join(opt["dir"], "*.xlsx")))
        archivos = [a for a in archivos if not os.path.basename(a).startswith("~$")]
        if not archivos:
            raise CommandError("No se indicó archivo ni --dir con .xlsx.")

        if dry:
            self.stdout.write(self.style.WARNING("\n⚠ DRY-RUN — no se escribe nada\n"))
        self.stdout.write(f"📚 UA: {ua}  |  archivos: {len(archivos)}\n")

        resumen = []          # filas del plan hoja→lab
        codigos_vistos = {}   # codigo → conteo (para desduplicar en el lote)
        total_creados = total_saltados = total_dup = 0
        ruteadas = {}         # lab destino → nº de filas movidas por --map-filas

        with transaction.atomic():
            if opt["replace_ua"]:
                qs = Equipo.objects.filter(unidad_academica=ua)
                n = qs.count()
                self.stdout.write(self.style.WARNING(
                    f"🗑  --replace-ua: eliminando {n} equipos de {ua.nombre}"))
                if not dry:
                    qs.delete()

            # codigo_activo es único GLOBAL: sembramos TODOS los códigos ya
            # presentes para desduplicar contra la BD y no sólo dentro del lote.
            # Incluye los de esta misma UA: al importar archivo por archivo (una
            # invocación por Excel) los cargados en la corrida anterior también
            # deben contar. Con --replace-ua ya se borraron justo arriba.
            for c in Equipo.objects.values_list("codigo_activo", flat=True):
                codigos_vistos[c] = codigos_vistos.get(c, 0) + 1

            for path in archivos:
                fname = os.path.basename(path)
                wb = openpyxl.load_workbook(path, data_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    hr, cols, lab_hint = self._detectar(ws)
                    if hr is None:
                        resumen.append((fname, sn, "—", "SIN CABECERA", 0))
                        continue

                    lab, metodo, destino = self._resolver_lab(
                        Laboratorio, ua, lab_hint, sn, fname, overrides, dry)

                    creados = 0
                    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                        cod = self._cell(row, cols, "cod")
                        nom = self._cell(row, cols, "nombre")
                        if not cod and not nom:
                            continue
                        if not cod:
                            # fila sin código: sólo la tomamos si tiene nombre real
                            cod = ""
                        # Desduplicar código dentro del lote
                        codigo = cod
                        if codigo:
                            if codigo in codigos_vistos:
                                codigos_vistos[codigo] += 1
                                codigo = f"{cod}/dup{codigos_vistos[cod]}"
                                total_dup += 1
                            else:
                                codigos_vistos[codigo] = 1
                        else:
                            # sin código: generar uno estable por UA+hoja+contador
                            codigos_vistos.setdefault("__SINCOD__", 0)
                            codigos_vistos["__SINCOD__"] += 1
                            codigo = f"SINCOD-{ua.pk}-{codigos_vistos['__SINCOD__']}"

                        estatus, estado_raw = parse_estado(self._cell(row, cols, "estado"))
                        espec = {}
                        marca = self._cell(row, cols, "marca")
                        anio = parse_anio(self._cell(row, cols, "anio"))
                        detalle = self._cell(row, cols, "espec")
                        func = self._cell(row, cols, "func")
                        if marca:
                            espec["marca_modelo"] = marca
                        anio_crudo = self._cell(row, cols, "anio")
                        if anio:
                            espec["anio_adquisicion"] = anio
                        elif anio_crudo:
                            espec["nota_adquisicion"] = anio_crudo[:200]
                        fecha_adq = parse_fecha_completa(
                            row[cols["anio"]] if "anio" in cols and cols["anio"] < len(row) else None
                        )
                        if fecha_adq:
                            espec["fecha_adquisicion"] = fecha_adq
                        if detalle:
                            espec["especificaciones"] = detalle[:8000]
                        if func:
                            espec["funcionalidad"] = func[:8000]

                        ubic = self._cell(row, cols, "ubic")
                        lab_fila = resolver_por_ubicacion(ubic, lab) if rutas_fila else lab
                        if lab_fila is not lab:
                            ruteadas[getattr(lab_fila, "nombre", "?")] = (
                                ruteadas.get(getattr(lab_fila, "nombre", "?"), 0) + 1)

                        if not dry:
                            e = Equipo(
                                nombre=(nom or f"EQUIPO {codigo}")[:150],
                                codigo_activo=codigo[:50],
                                laboratorio=lab_fila,
                                unidad_academica=ua,
                                cantidad_total=1,
                                cantidad_buena=1 if estatus == "bueno" else 0,
                                cantidad_regular=1 if estatus == "regular" else 0,
                                cantidad_mala=1 if estatus == "malo" else 0,
                                estatus_general=estatus,
                                ubicacion_sala=ubic[:255],
                                foto_url=primer_url(
                                    row[cols["foto1"]] if "foto1" in cols and cols["foto1"] < len(row) else None,
                                    row[cols["foto2"]] if "foto2" in cols and cols["foto2"] < len(row) else None,
                                ) or None,
                                especificaciones=espec,
                                observaciones=estado_raw if estatus == "malo" else "",
                            )
                            e.save()
                        creados += 1
                        if verbose and dry and creados <= 2:
                            self.stdout.write(f"     · {codigo:12} {nom[:34]:34} [{estatus}]")

                    total_creados += creados
                    resumen.append((fname, sn, f"{destino} · {metodo}",
                                    "OK" if lab or dry else "?", creados))

            if dry:
                transaction.set_rollback(True)

        # ── Reporte ──────────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 100)
        self.stdout.write(self.style.SUCCESS(
            f"PLAN DE IMPORTACIÓN {'(DRY-RUN)' if dry else '(APLICADO)'} — {ua.nombre}"))
        self.stdout.write("═" * 100)
        self.stdout.write(f"{'ARCHIVO':40} {'HOJA':26} {'→ LAB DESTINO · método':44} {'#eq':>4}")
        if ruteadas:
            self.stdout.write("─" * 100)
            self.stdout.write("RUTEO POR FILA (--map-filas):")
            for nombre, n in sorted(ruteadas.items(), key=lambda x: -x[1]):
                self.stdout.write(f"     {n:>4}  → {nombre}")
        self.stdout.write("─" * 100)
        for fname, sn, destino, estado, n in resumen:
            f = (fname[:38] + "…") if len(fname) > 39 else fname
            self.stdout.write(f"{f:40} {sn[:25]:26} {destino[:43]:44} {n:>4}  {estado}")
        self.stdout.write("─" * 100)
        self.stdout.write(self.style.SUCCESS(
            f"TOTAL equipos: {total_creados}  |  duplicados renombrados: {total_dup}"))
        self.stdout.write("═" * 100 + "\n")
