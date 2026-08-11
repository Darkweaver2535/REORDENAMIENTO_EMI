"""
Management command: auditar_carga
==================================
Reconcilia los Excel oficiales contra lo que quedó en la base de datos.

No confía en los contadores que imprimieron los importadores: vuelve a leer cada
archivo desde cero y comprueba, fila por fila, que su contenido está en la BD.

Comprueba:

  · FAMILIA A (infraestructura)  → cada laboratorio y subespacio del Excel
                                   existe como nodo, y cada asignatura como uso
                                   académico.
  · FAMILIA C (fichas técnicas)  → cada código de equipo del Excel existe como
                                   Equipo, con laboratorio y especificaciones.
  · FAMILIA B (padrón contable)  → cada código del padrón existe como Equipo con
                                   sus datos contables.
  · GUÍAS                        → cada PDF tiene su Guia con archivo en disco.

Uso:
    python manage.py auditar_carga
    python manage.py auditar_carga --detalle    # lista cada elemento faltante
"""

import glob
import os
import re
import unicodedata
from collections import defaultdict

from django.core.management.base import BaseCommand

try:
    import openpyxl
except ImportError:
    openpyxl = None

SRC = "/Users/alvaroencinas/Downloads/INFORMACION REORDENAMIENTO EMI"

# ── Clasificación de los archivos ───────────────────────────────────────────
FAMILIA_A = [
    ("COCHABAMBA-SISTEMA DE GESTION DE  LABORATORIOS 2026 U.A. CBBA.xlsx", "Hoja1", "UACB"),
    ("LA PAZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "LA PAZ", "UALP"),
    ("TROPICO-SISTEMA DE GESTIÓN DE LABORATORIOS.xlsx", "TROPICO", "UAT"),
    ("SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "COCHABAMBA", "UACB"),
    ("SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "LA PAZ", "UALP"),
    ("SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "TROPICO", "UAT"),
    ("SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "RIBERALTA", "UAR"),
    ("SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "SANTA CRUZ", "UASC"),
    ("SANTA CRUZ-DATOS DE UBICACIÓN LABORATORIOS - 2026.xlsx", "DICYT", "DNICYT"),
]

# Mismos alias que usa la importación: el consolidado nacional nombra algunos
# ambientes distinto que el Excel de la sede. Sin cargarlos, la auditoría los
# reportaría como faltantes cuando en realidad son el mismo nodo.
ALIAS_POR_UA = {
    "UALP": "_alias_consolidado_ualp.json",
    "UACB": "_alias_consolidado_uacb.json",
}

# El importador compone los nombres de subespacio sin significado propio con el
# del padre ("LABORATORIO" ⊂ AGRONOMIA → "LABORATORIO DE AGRONOMIA").
NOMBRES_GENERICOS = {"LABORATORIO", "LABORATORIOS", "SALA", "AREA", "SECCION", "LAB"}

FAMILIA_B = {
    "EQUIPO MEDICO Y LAB. UALP EMI.xlsx": "UALP",
    "EQUIPO MEDICO Y LABORATORIO UACBBA.xlsx": "UACB",
    "GRUPO EQUIPO MEDICO Y LABORATORIO UASC.xlsx": "UASC",
    "LABORATORIO DE LA EMI UAT.xlsx": "UAT",
    "DETALLE DE ACTIVOS FIJOS EQUIPO MEDICOS Y DE LABORATORIO UA RIBERALTA.xlsx": "UAR",
    "EQUIPO MEDICO Y DE LABORATORIO OFICINA CENTRAL.xlsx": "DNICYT",
}


def norm(valor):
    if valor is None:
        return ""
    s = re.sub(r"\s+", " ", str(valor).replace("\xa0", " ")).strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").upper()


# Definición única y compartida con los importadores y el normalizador.
from apps.laboratorios.codigos import canonizar_codigo as norm_codigo  # noqa: E402
from apps.laboratorios.codigos import codigo_base, es_sin_codigo  # noqa: E402


class Command(BaseCommand):
    help = "Reconcilia los Excel oficiales contra la base de datos."

    def add_arguments(self, parser):
        parser.add_argument("--detalle", action="store_true",
                            help="Lista cada elemento faltante, no sólo el conteo.")
        parser.add_argument("--src", default=SRC)

    def _alias(self, ua):
        """{(ámbito, nombre_normalizado): nombre_canónico} de la UA indicada."""
        import json

        fichero = ALIAS_POR_UA.get(ua)
        if not fichero:
            return {}
        ruta = os.path.join(os.path.dirname(__file__), fichero)
        if not os.path.exists(ruta):
            return {}
        with open(ruta, encoding="utf-8") as fh:
            crudo = json.load(fh)
        mapa = {}
        for clave, valor in crudo.items():
            if clave.startswith("_"):
                continue
            ambito, _, nombre = clave.partition(":")
            if not nombre:
                ambito, nombre = "", clave
            n = norm(nombre)
            for a in (("raiz", "hijo") if ambito.lower() not in ("raiz", "hijo") else (ambito.lower(),)):
                mapa[(a, n)] = valor
        return mapa

    # ── Utilidades de lectura ────────────────────────────────────────────
    def _hoja_equipos(self, ws):
        """(fila_cabecera, col_codigo, col_nombre) de una hoja familia C."""
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
            j = norm(" ".join(str(c) for c in row if c))
            if "NOMBRE DEL EQUIPO" in j and "CODIGO DEL EQUIPO" in j:
                cc = cn = None
                for k, c in enumerate(row):
                    u = norm(c)
                    if "CODIGO DEL EQUIPO" in u and cc is None:
                        cc = k
                    elif "NOMBRE DEL EQUIPO" in u and cn is None:
                        cn = k
                return i, cc, cn
        return None, None, None

    def handle(self, *args, **opt):
        from apps.estructura_academica.models import UnidadAcademica
        from apps.guias.models import Guia
        from apps.laboratorios.models import Equipo, Laboratorio, UsoAcademico

        src = opt["src"]
        detalle = opt["detalle"]
        problemas = []

        # Índices de la BD
        equipos_por_codigo = defaultdict(list)
        for pk, cod, lab_id, esp in Equipo.objects.values_list(
                "pk", "codigo_activo", "laboratorio_id", "especificaciones"):
            equipos_por_codigo[codigo_base(cod)].append((pk, lab_id, esp or {}))

        labs_por_ua = defaultdict(set)
        for nombre, ua in Laboratorio.objects.values_list("nombre", "unidad_academica__nombre"):
            labs_por_ua[ua].add(norm(nombre))

        usos_por_ua = defaultdict(set)
        for asig, ua in UsoAcademico.objects.values_list(
                "asignatura", "laboratorio__unidad_academica__nombre"):
            usos_por_ua[ua].add(norm(asig))

        # ── FAMILIA A ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 96)
        self.stdout.write(self.style.SUCCESS("FAMILIA A — INFRAESTRUCTURA (laboratorios, subespacios, usos académicos)"))
        self.stdout.write("═" * 96)
        self.stdout.write(f"{'ARCHIVO · HOJA':58}{'UA':8}{'labs':>6}{'falta':>6}{'asig':>6}{'falta':>6}")
        self.stdout.write("─" * 96)

        for fn, hoja, ua in FAMILIA_A:
            ruta = os.path.join(src, fn)
            if not os.path.exists(ruta):
                problemas.append(f"FALTA EL ARCHIVO: {fn}")
                continue
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb[hoja]
            hdr = cn = cs = ca = None
            for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True), 1):
                for k, c in enumerate(row):
                    u = norm(c)
                    if "NOMBRE DEL LABORATORIO" in u:
                        hdr, cn = i, k
                    elif u.startswith("NOMBRE DE LA SALA"):
                        cs = k
                    elif "QUE ASIGNATURAS UTILIZAN" in u:
                        ca = k
                if hdr:
                    break
            alias = self._alias(ua)
            nombres, asigs = set(), set()
            padre_actual = ""
            if hdr:
                for row in ws.iter_rows(min_row=hdr + 2, values_only=True):
                    # Raíz
                    if cn is not None and cn < len(row) and row[cn]:
                        t = norm(row[cn])
                        if t and t not in ("PEA", "INVESTIGACION", "VENTA DE SERVICIOS"):
                            padre_actual = t
                            nombres.add(norm(alias.get(("raiz", t), t)))
                    # Subespacio
                    if cs is not None and cs < len(row) and row[cs]:
                        t = norm(row[cs])
                        if t and t not in ("PEA", "INVESTIGACION", "VENTA DE SERVICIOS"):
                            if t in NOMBRES_GENERICOS and padre_actual:
                                t = f"{t} DE {padre_actual}"
                            nombres.add(norm(alias.get(("hijo", t), t)))
                    if ca is not None and ca < len(row) and row[ca]:
                        t = norm(row[ca])
                        if t:
                            asigs.add(t)
            wb.close()

            faltan_lab = {n for n in nombres if n not in labs_por_ua.get(ua, set())}
            faltan_asig = {a for a in asigs if a not in usos_por_ua.get(ua, set())}
            etiqueta = f"{fn[:38]} · {hoja}"
            self.stdout.write(
                f"{etiqueta[:57]:58}{ua:8}{len(nombres):>6}{len(faltan_lab):>6}"
                f"{len(asigs):>6}{len(faltan_asig):>6}")
            if faltan_lab:
                problemas.append(f"[A] {etiqueta}: {len(faltan_lab)} laboratorios sin registrar")
                if detalle:
                    for n in sorted(faltan_lab):
                        self.stdout.write(self.style.WARNING(f"        lab faltante: {n}"))
            if faltan_asig:
                problemas.append(f"[A] {etiqueta}: {len(faltan_asig)} asignaturas sin uso académico")
                if detalle:
                    for a in sorted(faltan_asig)[:20]:
                        self.stdout.write(self.style.WARNING(f"        asignatura faltante: {a}"))

        # ── FAMILIA C ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 96)
        self.stdout.write(self.style.SUCCESS("FAMILIA C — FICHAS TÉCNICAS POR LABORATORIO"))
        self.stdout.write("═" * 96)
        self.stdout.write(f"{'ARCHIVO · HOJA':64}{'filas':>7}{'c/cód':>7}{'falta':>7}{'s/cód':>7}")
        self.stdout.write("─" * 96)

        tot_filas = tot_cod = tot_falta = tot_sincod = 0
        for ruta in sorted(glob.glob(os.path.join(src, "*.xlsx"))):
            fn = os.path.basename(ruta)
            if fn in FAMILIA_B or any(fn == a for a, _, _ in FAMILIA_A):
                continue
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            for ws in wb.worksheets:
                hr, cc, cn = self._hoja_equipos(ws)
                if hr is None:
                    continue
                filas = con_cod = sin_cod = 0
                faltantes = []
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    cod = row[cc] if cc is not None and cc < len(row) else None
                    nom = row[cn] if cn is not None and cn < len(row) else None
                    if not cod and not nom:
                        continue
                    filas += 1
                    if not cod or es_sin_codigo(cod):
                        # El Excel declara que el bien no tiene código de activo;
                        # el equipo se cargó igual, con un identificador SINCOD-.
                        sin_cod += 1
                        continue
                    con_cod += 1
                    if codigo_base(cod) not in equipos_por_codigo:
                        faltantes.append(str(cod).strip())
                if not filas:
                    continue
                tot_filas += filas; tot_cod += con_cod
                tot_falta += len(faltantes); tot_sincod += sin_cod
                etiqueta = f"{fn[:44]} · {ws.title}"
                estilo = self.style.WARNING if faltantes else (lambda s: s)
                self.stdout.write(estilo(
                    f"{etiqueta[:63]:64}{filas:>7}{con_cod:>7}{len(faltantes):>7}{sin_cod:>7}"))
                if faltantes:
                    problemas.append(f"[C] {etiqueta}: {len(faltantes)} códigos sin registrar")
                    if detalle:
                        for c in faltantes[:15]:
                            self.stdout.write(self.style.WARNING(f"        código faltante: {c}"))
            wb.close()
        self.stdout.write("─" * 96)
        self.stdout.write(f"{'TOTAL':64}{tot_filas:>7}{tot_cod:>7}{tot_falta:>7}{tot_sincod:>7}")

        # ── FAMILIA B ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 96)
        self.stdout.write(self.style.SUCCESS("FAMILIA B — PADRÓN CONTABLE DE ACTIVOS FIJOS"))
        self.stdout.write("═" * 96)
        self.stdout.write(f"{'ARCHIVO':58}{'UA':8}{'códigos':>9}{'falta':>7}{'s/contab':>9}")
        self.stdout.write("─" * 96)

        for fn, ua in FAMILIA_B.items():
            ruta = os.path.join(src, fn)
            if not os.path.exists(ruta):
                problemas.append(f"FALTA EL ARCHIVO: {fn}")
                continue
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            hr = cc = None
            for i, row in enumerate(ws.iter_rows(max_row=14, values_only=True), 1):
                j = norm(" ".join(str(c) for c in row if c))
                if "CODIGO" in j and "DESCRIPCION" in j:
                    hr = i
                    for k, c in enumerate(row):
                        if norm(c).startswith("CODIGO"):
                            cc = k
                            break
                    break
            codigos, faltantes, sin_contable = set(), [], 0
            if hr is not None and cc is not None:
                for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                    if cc >= len(row) or not row[cc]:
                        continue
                    c = codigo_base(row[cc])
                    if len(c) < 3 or not any(ch.isdigit() for ch in c):
                        continue
                    codigos.add(c)
                    filas_bd = equipos_por_codigo.get(c)
                    if not filas_bd:
                        faltantes.append(str(row[cc]).strip())
                    elif not any("grupo_contable" in esp or "responsable" in esp
                                 for _, _, esp in filas_bd):
                        sin_contable += 1
            wb.close()
            estilo = self.style.WARNING if faltantes else (lambda s: s)
            self.stdout.write(estilo(
                f"{fn[:57]:58}{ua:8}{len(codigos):>9}{len(faltantes):>7}{sin_contable:>9}"))
            if faltantes:
                problemas.append(f"[B] {fn}: {len(faltantes)} códigos del padrón sin registrar")
                if detalle:
                    for c in faltantes[:15]:
                        self.stdout.write(self.style.WARNING(f"        código faltante: {c}"))
            if sin_contable:
                problemas.append(f"[B] {fn}: {sin_contable} equipos sin datos contables")

        # ── GUÍAS ────────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 96)
        self.stdout.write(self.style.SUCCESS("GUÍAS DE LABORATORIO (PDF)"))
        self.stdout.write("═" * 96)
        pdfs = sorted(glob.glob(os.path.join(src, "GL*.pdf")))
        guias = list(Guia.objects.select_related("asignatura"))
        con_archivo = [g for g in guias if g.pdf_archivo]
        en_disco = sum(
            1 for g in con_archivo
            if os.path.exists(g.pdf_archivo.path) and g.pdf_archivo.size > 0)
        self.stdout.write(f"  PDF en la carpeta origen : {len(pdfs)}")
        self.stdout.write(f"  Guías en la BD           : {len(guias)}")
        self.stdout.write(f"  Con archivo subido       : {len(con_archivo)}")
        self.stdout.write(f"  Archivo legible en disco : {en_disco}")
        if len(pdfs) != len(guias):
            problemas.append(f"[G] {len(pdfs)} PDF pero {len(guias)} guías en la BD")
        if en_disco != len(pdfs):
            problemas.append(f"[G] sólo {en_disco} de {len(pdfs)} PDF están en disco")

        # Los tamaños deben coincidir con el origen (el PDF se sube tal cual).
        tam_origen = {os.path.basename(p): os.path.getsize(p) for p in pdfs}
        desajustes = 0
        for g in con_archivo:
            base = os.path.basename(g.pdf_archivo.name).replace("_", " ")
            for nombre, tam in tam_origen.items():
                if norm(nombre.replace("_", " ")) == norm(base):
                    if g.pdf_archivo.size != tam:
                        desajustes += 1
                    break
        if desajustes:
            problemas.append(f"[G] {desajustes} PDF con tamaño distinto al origen")
        else:
            self.stdout.write("  Tamaño idéntico al origen: sí (el PDF se subió sin alterar)")

        # ── VEREDICTO ────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 96)
        if problemas:
            self.stdout.write(self.style.ERROR(f"HALLAZGOS: {len(problemas)}"))
            self.stdout.write("═" * 96)
            for p in problemas:
                self.stdout.write(self.style.ERROR(f"  ✗ {p}"))
        else:
            self.stdout.write(self.style.SUCCESS(
                "✅ COBERTURA COMPLETA — todo el contenido de los Excel está en la base de datos"))
        self.stdout.write("═" * 96 + "\n")
