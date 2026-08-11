"""
Management command: import_activos
===================================
Importa el inventario CONTABLE de activos fijos (grupo "EQUIPO MEDICO Y DE
LABORATORIO") desde los Excel de Activos Fijos de cada unidad académica.

Complementa a `import_equipos`, que carga las fichas técnicas por laboratorio:

  · `import_equipos`  → qué equipo hay en cada laboratorio, con especificaciones,
                        marca/modelo, fotos y estado físico. Es la fuente de la
                        UBICACIÓN.
  · `import_activos`  → el padrón contable completo de la sede. Es la fuente del
                        RESPONSABLE, grupo contable, auxiliar, costo y fechas, y
                        aporta los activos que ninguna ficha de laboratorio
                        reportó.

Ambas fuentes comparten `codigo_activo`, así que el comando:

  · ENRIQUECE el equipo existente con los datos contables (no pisa el nombre, el
    laboratorio ni el estado físico, que vienen de la ficha, más reciente).
  · CREA el equipo cuando el código no existe todavía.

Ubicación de los activos nuevos: la columna OFICINA se traduce a un laboratorio
con `--map-oficinas`. Las oficinas que no son laboratorios (secciones de activos
fijos, depósitos, unidades administrativas) quedan con `laboratorio = NULL` y su
nombre en `ubicacion_sala`: son inventario real pendiente de asignar, insumo
natural del módulo de reordenamiento.

Formatos de cabecera soportados (se detecta por texto, no por posición):
  NRO|CÓDIGO|DESCRIPCIÓN DEL BIEN|OBSERVACIONES|COSTO|FECHA|ESTADO|VIDA ÚTIL|
  GRUPO CONTABLE|AUXILIAR|OFICINA|RESPONSABLE|CARGO|CARNET      (UALP/UACB/UASC/DNICYT)
  ... sin COSTO ni FECHA                                        (UAT)
  UNIDAD|CÓDIGO|DESCRIPCIÓN|RESPONSABLE|GRUPO|AUXILIAR|VIDA|OFICINA|FECHAS|ESTADO  (UAR)

Uso:
    python manage.py import_activos "ruta/inventario.xlsx" --unidad-academica UACB \\
        --map-oficinas apps/laboratorios/management/commands/_oficinas_uacb.json --dry-run
"""

import json
import os
import re
import unicodedata
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None


def normalizar(valor) -> str:
    if valor is None:
        return ""
    s = str(valor).replace("\xa0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").upper()


def limpiar(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return re.sub(r"\s+", " ", str(valor).replace("\xa0", " ").replace("\n", " ")).strip()


from apps.laboratorios.codigos import canonizar_codigo as norm_codigo  # noqa: E402


def codigo_valido(codigo: str) -> bool:
    """Descarta encabezados repetidos, totales y celdas basura, no formatos raros."""
    return len(codigo) >= 3 and any(c.isdigit() for c in codigo)


def parse_estado(raw: str) -> str:
    n = normalizar(raw)
    if n.startswith("BUEN") or n.startswith("NUEVO"):
        return "bueno"
    if n.startswith("MAL") or n.startswith("INSERV") or n.startswith("OBSOLET"):
        return "malo"
    return "regular"


CAMPOS = {
    "cod": ("CODIGO",),
    "desc": ("DESCRIPCION DEL BIEN", "DESCRIPCION"),
    "obs": ("OBSERVACIONES",),
    "costo": ("COSTO HISTORICO",),
    "fecha": ("FECHA HISTORICO", "FECHA INC"),
    "estado": ("ESTADO",),
    "vida": ("VIDA UTIL CONSUMIDA", "VIDA UTIL"),
    "grupo": ("GRUPO CONTABLE",),
    "aux": ("AUXILIAR",),
    "ofi": ("OFICINA",),
    "resp": ("RESPONSABLE",),
    "cargo": ("CARGO",),
}


class Command(BaseCommand):
    help = "Importa el padrón contable de activos fijos y lo cruza con los equipos."

    def add_arguments(self, parser):
        parser.add_argument("archivo", help="Ruta al .xlsx del inventario.")
        parser.add_argument("--unidad-academica", required=True, metavar="UA")
        parser.add_argument("--map-oficinas", default=None,
                            help="JSON {OFICINA: nombre_lab_destino}. Las oficinas "
                                 "ausentes o con valor null dejan laboratorio=NULL.")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--verbose", action="store_true")

    def _resolver_ua(self, UnidadAcademica, ref):
        ref_n = normalizar(ref)
        for ua in UnidadAcademica.objects.all():
            if normalizar(ua.nombre) == ref_n or normalizar(ua.codigo) == ref_n:
                return ua
        raise CommandError(f"UA '{ref}' no encontrada.")

    def _detectar(self, ws):
        """Devuelve (fila_cabecera, {campo: indice}) buscando por texto."""
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
            joined = normalizar(" ".join(str(c) for c in row if c))
            if "CODIGO" not in joined or "DESCRIPCION" not in joined:
                continue
            cols = {}
            for k, celda in enumerate(row):
                u = normalizar(celda)
                if not u:
                    continue
                for campo, prefijos in CAMPOS.items():
                    if campo in cols:
                        continue
                    if any(u.startswith(p) for p in prefijos):
                        cols[campo] = k
                        break
            if "cod" in cols and "desc" in cols:
                return i, cols
        return None, None

    def handle(self, *args, **opt):
        if openpyxl is None:
            raise CommandError("openpyxl no instalado")

        from apps.estructura_academica.models import UnidadAcademica
        from apps.laboratorios.models import Equipo, Laboratorio

        ua = self._resolver_ua(UnidadAcademica, opt["unidad_academica"])
        dry, verbose = opt["dry_run"], opt["verbose"]

        oficinas = {}
        if opt["map_oficinas"]:
            with open(opt["map_oficinas"], encoding="utf-8") as fh:
                oficinas = {normalizar(k): v
                            for k, v in json.load(fh).items() if not k.startswith("_")}

        # Índice de laboratorios hoja de la UA, por nombre normalizado.
        hojas = {normalizar(l.nombre): l
                 for l in Laboratorio.objects.filter(unidad_academica=ua) if l.es_hoja()}

        # Índice de los equipos ya cargados por las fichas técnicas.
        por_codigo = {norm_codigo(c): pk for pk, c in
                      Equipo.objects.values_list("pk", "codigo_activo")}

        wb = openpyxl.load_workbook(opt["archivo"], read_only=True, data_only=True)
        ws = wb.worksheets[0]
        hr, cols = self._detectar(ws)
        if hr is None:
            raise CommandError("No se encontró la cabecera (CÓDIGO / DESCRIPCIÓN).")

        def cell(row, key):
            k = cols.get(key)
            return limpiar(row[k]) if k is not None and k < len(row) else ""

        enriquecidos = creados = sin_lab = saltados = 0
        ubic_faltantes = {}
        destinos = {}
        nuevos_codigos = set()

        if dry:
            self.stdout.write(self.style.WARNING("\n⚠ DRY-RUN — no se escribe nada\n"))
        self.stdout.write(f"📚 UA: {ua.nombre}  |  cabecera fila {hr}  |  columnas: {sorted(cols)}\n")

        with transaction.atomic():
            for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                cod_raw = cell(row, "cod")
                desc = cell(row, "desc")
                if not cod_raw or not desc:
                    continue
                codigo = norm_codigo(cod_raw)
                if not codigo_valido(codigo):
                    saltados += 1
                    continue

                contable = {
                    "grupo_contable": cell(row, "grupo"),
                    "auxiliar": cell(row, "aux"),
                    "responsable": cell(row, "resp"),
                    "cargo_responsable": cell(row, "cargo"),
                    "oficina_contable": cell(row, "ofi"),
                    "costo_historico": cell(row, "costo"),
                    "fecha_historico": cell(row, "fecha"),
                    "vida_util_consumida": cell(row, "vida"),
                    "codigo_contable": cod_raw,
                }
                contable = {k: v for k, v in contable.items() if v}

                pk = por_codigo.get(codigo)
                if pk is not None:
                    # Ya existe por ficha técnica → sólo se agregan los datos contables.
                    if not dry:
                        eq = Equipo.objects.get(pk=pk)
                        espec = dict(eq.especificaciones or {})
                        espec.update(contable)
                        eq.especificaciones = espec
                        campos = ["especificaciones"]
                        obs = cell(row, "obs")
                        if obs and obs not in (eq.observaciones or ""):
                            eq.observaciones = (
                                f"{eq.observaciones}\n{obs}".strip()
                                if eq.observaciones else obs
                            )
                            campos.append("observaciones")
                        eq.save(update_fields=campos)
                    enriquecidos += 1
                    continue

                if codigo in nuevos_codigos:
                    saltados += 1
                    continue
                nuevos_codigos.add(codigo)

                # Activo que ninguna ficha de laboratorio reportó → alta.
                ofi = cell(row, "ofi")
                destino_nombre = oficinas.get(normalizar(ofi))
                lab = hojas.get(normalizar(destino_nombre)) if destino_nombre else None
                if destino_nombre and lab is None:
                    ubic_faltantes[destino_nombre] = ubic_faltantes.get(destino_nombre, 0) + 1
                if lab is None:
                    sin_lab += 1
                destinos[lab.nombre if lab else f"— sin laboratorio ({ofi or 's/oficina'})"] = (
                    destinos.get(lab.nombre if lab else f"— sin laboratorio ({ofi or 's/oficina'})", 0) + 1
                )

                estatus = parse_estado(cell(row, "estado"))
                if not dry:
                    Equipo.objects.create(
                        nombre=desc[:150],
                        codigo_activo=codigo[:50],
                        laboratorio=lab,
                        unidad_academica=ua,
                        cantidad_total=1,
                        cantidad_buena=1 if estatus == "bueno" else 0,
                        cantidad_regular=1 if estatus == "regular" else 0,
                        cantidad_mala=1 if estatus == "malo" else 0,
                        estatus_general=estatus,
                        ubicacion_sala=ofi[:100],
                        observaciones=cell(row, "obs"),
                        especificaciones=contable,
                    )
                creados += 1
                if verbose and creados <= 5:
                    self.stdout.write(f"     + {codigo:12} {desc[:46]:46} → {lab or 'NULL'}")

            if dry:
                transaction.set_rollback(True)

        wb.close()

        self.stdout.write("\n" + "═" * 92)
        self.stdout.write(self.style.SUCCESS(
            f"INVENTARIO CONTABLE {'(DRY-RUN)' if dry else '(APLICADO)'} — {ua.nombre}"))
        self.stdout.write("═" * 92)
        self.stdout.write("DESTINO DE LOS ACTIVOS NUEVOS:")
        for nombre, n in sorted(destinos.items(), key=lambda x: -x[1]):
            self.stdout.write(f"     {n:>5}  {nombre}")
        if ubic_faltantes:
            self.stdout.write(self.style.WARNING("\n⚠ Destinos de --map-oficinas que no existen como nodo hoja:"))
            for nombre, n in ubic_faltantes.items():
                self.stdout.write(self.style.WARNING(f"     {n:>5}  «{nombre}»"))
        self.stdout.write("─" * 92)
        self.stdout.write(self.style.SUCCESS(
            f"enriquecidos (ya venían de la ficha): {enriquecidos}\n"
            f"creados (sólo estaban en el padrón):  {creados}   ·  de ellos sin laboratorio: {sin_lab}\n"
            f"filas descartadas (código inválido/duplicado): {saltados}"))
        self.stdout.write("═" * 92 + "\n")
